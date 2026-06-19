
import math
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats.diagnostic import het_breuschpagan, het_white
from statsmodels.stats.outliers_influence import variance_inflation_factor, OLSInfluence
from statsmodels.stats.stattools import durbin_watson
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="STATCAL ONLINE - IDX Banking Financial Data Analyzer",
    page_icon="🏦",
    layout="wide",
)


# =====================================================
# CONSTANTS
# =====================================================

APP_NAME = "STATCAL ONLINE"
APP_TITLE = "IDX Banking Financial Data Analyzer"
APP_UPDATED = "Last updated on June 19, 2026"
WEBSITE_URL = "https://statcal.com/"
STATCAL_ONLINE_URL = "https://statcal.com/statcal%20online.html"
TRAINING_DATA_URL = "https://drive.google.com/drive/folders/1s273Ad5FUElhzd5G16jWSBxbOtforzRR?usp=sharing"
SAMPLE_DATA_PATH = Path("data idx perbankan.xlsx")
LOGO_PATH = Path("logo_statcal.png")
BRAND_COLOR = "1F4E79"

FINANCIAL_URL_COL = "Financial Statement URL"

COLOR_PALETTES = {
    "Scopus Blue Orange": ["#1F4E79", "#E97132", "#70AD47", "#FFC000", "#7030A0", "#00A6A6", "#C00000", "#595959"],
    "Nature Publication": ["#1B7837", "#762A83", "#5AAE61", "#9970AB", "#A6DBA0", "#C2A5CF", "#00441B", "#40004B"],
    "Finance Professional": ["#0B3C49", "#3282B8", "#BBE1FA", "#F9A03F", "#D1495B", "#00798C", "#30638E", "#003D5B"],
    "Economics Journal": ["#003F5C", "#BC5090", "#FFA600", "#58508D", "#FF6361", "#2F4B7C", "#A05195", "#D45087"],
    "Black Gray Academic": ["#000000", "#404040", "#666666", "#808080", "#A6A6A6", "#BFBFBF", "#595959", "#262626"],
    "High Contrast Q1": ["#2166AC", "#B2182B", "#4D9221", "#762A83", "#D6604D", "#1B7837", "#4393C3", "#9970AB"],
}

THEMES = {
    "White Publication": {
        "figure_facecolor": "white",
        "axes_facecolor": "white",
        "text_color": "#111111",
        "grid_color": "#D9D9D9",
        "spine_color": "#222222",
    },
    "Light Gray Editorial": {
        "figure_facecolor": "#F7F7F7",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#111111",
        "grid_color": "#D0D0D0",
        "spine_color": "#333333",
    },
    "Warm Ivory Journal": {
        "figure_facecolor": "#FBF7EF",
        "axes_facecolor": "#FFFDF8",
        "text_color": "#1F1F1F",
        "grid_color": "#DDD4C4",
        "spine_color": "#3A3A3A",
    },
    "Cool Blue Scientific": {
        "figure_facecolor": "#F3F7FB",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#0B1F33",
        "grid_color": "#C8D6E5",
        "spine_color": "#1F4E79",
    },
    "Dark Navy Presentation": {
        "figure_facecolor": "#0B1320",
        "axes_facecolor": "#111C2E",
        "text_color": "#FFFFFF",
        "grid_color": "#3B4A5F",
        "spine_color": "#B8C7D9",
    },
}

MARKERS = ["o", "s", "^", "D", "P", "X", "v", "<", ">", "*"]


# =====================================================
# GENERAL HELPERS
# =====================================================

def safe_streamlit_image(image_path: Path, width: int = 220) -> None:
    try:
        st.image(str(image_path), width=width)
    except Exception:
        st.markdown("### STATCAL")


@st.cache_data(ttl=3600, max_entries=20)
def read_excel_file(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)


@st.cache_data(ttl=3600, max_entries=10)
def get_excel_sheet_names(file_bytes: bytes) -> List[str]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes))
    return excel_file.sheet_names


@st.cache_data(ttl=3600, max_entries=20)
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(col)).strip() for col in df.columns]
    df = df.dropna(how="all")
    unnamed_cols = [col for col in df.columns if str(col).lower().startswith("unnamed")]
    for col in unnamed_cols:
        if df[col].isna().all():
            df = df.drop(columns=[col])
    return df.reset_index(drop=True)


def make_arrow_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    for col in safe_df.columns:
        if safe_df[col].dtype == "object":
            safe_df[col] = safe_df[col].astype(str).replace({"nan": "", "None": "", "NaT": ""})
    return safe_df


def to_numeric_series(series: pd.Series) -> pd.Series:
    """Convert a pandas Series to numeric values. Supports commas, %, and K/M/B/T suffixes."""
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")

    text = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("−", "-", regex=False)
        .str.replace("—", "", regex=False)
        .str.strip()
    )

    text = text.replace({
        "": pd.NA,
        "nan": pd.NA,
        "None": pd.NA,
        "NaT": pd.NA,
        "-": pd.NA,
        "N/A": pd.NA,
        "NA": pd.NA,
        "n/a": pd.NA,
        "na": pd.NA,
    })

    def convert_value(value):
        if pd.isna(value):
            return np.nan
        value = str(value).strip()
        if not value:
            return np.nan

        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1].strip()

        multiplier = 1.0
        last_char = value[-1:].lower()
        if last_char == "k":
            multiplier = 1_000.0
            value = value[:-1]
        elif last_char == "m":
            multiplier = 1_000_000.0
            value = value[:-1]
        elif last_char == "b":
            multiplier = 1_000_000_000.0
            value = value[:-1]
        elif last_char == "t":
            multiplier = 1_000_000_000_000.0
            value = value[:-1]

        value = value.replace("%", "").strip()
        try:
            number = float(value) * multiplier
            return -number if negative else number
        except Exception:
            return np.nan

    return text.map(convert_value)


@st.cache_data(ttl=3600, max_entries=20)
def detect_numeric_columns(df: pd.DataFrame, min_valid_ratio: float = 0.45) -> List[str]:
    numeric_cols = []
    for col in df.columns:
        s = df[col]
        non_null = s.notna().sum()
        if non_null == 0:
            continue
        numeric_s = to_numeric_series(s)
        valid_ratio = numeric_s.notna().sum() / max(non_null, 1)
        if valid_ratio >= min_valid_ratio:
            numeric_cols.append(col)
    return numeric_cols


def sorted_unique_values(series: pd.Series) -> List:
    values = series.dropna().unique().tolist()
    try:
        return sorted(values)
    except Exception:
        return sorted(values, key=lambda x: str(x).lower())


def preferred_option(options: List[str], candidates: List[str], fallback_index: int = 0) -> int:
    for candidate in candidates:
        for idx, option in enumerate(options):
            if str(option).strip().lower() == candidate.lower():
                return idx
    return fallback_index


def default_numeric_columns(columns: List[str]) -> List[str]:
    excluded = {"year", "tahun"}
    return [col for col in columns if str(col).strip().lower() not in excluded][:5]


def get_theme(theme_name: str) -> Dict[str, str]:
    return THEMES.get(theme_name, THEMES["White Publication"])


def get_palette_color_list(palette_name: str) -> List[str]:
    return COLOR_PALETTES.get(palette_name, COLOR_PALETTES["Scopus Blue Orange"])


def format_numeric_label(value: float, decimal_digits: int, compact: bool = False) -> str:
    if pd.isna(value) or not np.isfinite(value):
        return ""
    if compact:
        abs_value = abs(value)
        if abs_value >= 1_000_000_000_000:
            return f"{value / 1_000_000_000_000:.{decimal_digits}f}T"
        if abs_value >= 1_000_000_000:
            return f"{value / 1_000_000_000:.{decimal_digits}f}B"
        if abs_value >= 1_000_000:
            return f"{value / 1_000_000:.{decimal_digits}f}M"
        if abs_value >= 1_000:
            return f"{value / 1_000:.{decimal_digits}f}K"
    return f"{value:.{decimal_digits}f}"


def format_x_labels(values: List) -> List[str]:
    labels = []
    for value in values:
        if pd.isna(value):
            labels.append("Missing")
        elif isinstance(value, pd.Timestamp):
            labels.append(value.strftime("%Y-%m-%d"))
        elif isinstance(value, (np.integer, int)):
            labels.append(str(int(value)))
        else:
            labels.append(str(value))
    return labels


# =====================================================
# STATISTICS
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def compute_descriptive_statistics(df: pd.DataFrame, numeric_cols: List[str], decimal_digits: int) -> pd.DataFrame:
    rows = []
    for col in numeric_cols:
        s = to_numeric_series(df[col]).dropna()
        rows.append({
            "Variable": col,
            "N": int(s.count()),
            "Minimum": s.min() if len(s) else np.nan,
            "Maximum": s.max() if len(s) else np.nan,
            "Mean": s.mean() if len(s) else np.nan,
            "Median": s.median() if len(s) else np.nan,
            "Standard Deviation": s.std(ddof=1) if len(s) > 1 else np.nan,
            "Variance": s.var(ddof=1) if len(s) > 1 else np.nan,
            "Skewness": s.skew() if len(s) > 2 else np.nan,
            "Kurtosis": s.kurt() if len(s) > 3 else np.nan,
        })
    result = pd.DataFrame(rows)
    numeric_output_cols = [c for c in result.columns if c not in ["Variable", "N"]]
    for col in numeric_output_cols:
        result[col] = result[col].round(decimal_digits)
    return result


@st.cache_data(ttl=3600, max_entries=20)
def compute_grouped_descriptive_statistics(
    df: pd.DataFrame,
    numeric_cols: List[str],
    group_cols: List[str],
    decimal_digits: int,
) -> pd.DataFrame:
    if not group_cols:
        return compute_descriptive_statistics(df, numeric_cols, decimal_digits)

    rows = []
    for numeric_col in numeric_cols:
        temp = df[group_cols].copy()
        temp["__value__"] = to_numeric_series(df[numeric_col])
        grouped = temp.groupby(group_cols, dropna=False)["__value__"]
        stats_df = grouped.agg(
            N="count",
            Minimum="min",
            Maximum="max",
            Mean="mean",
            Median="median",
            **{"Standard Deviation": "std"},
        ).reset_index()
        stats_df.insert(len(group_cols), "Variable", numeric_col)
        rows.append(stats_df)

    result = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    for col in ["Minimum", "Maximum", "Mean", "Median", "Standard Deviation"]:
        if col in result.columns:
            result[col] = result[col].round(decimal_digits)
    return result


# =====================================================
# LINE CHART
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def build_mean_line_long_data(
    df: pd.DataFrame,
    x_col: str,
    numeric_cols: List[str],
    split_col: str,
    sort_x: bool,
) -> pd.DataFrame:
    rows = []
    for variable in numeric_cols:
        use_cols = [x_col, variable]
        use_split = split_col != "None" and split_col in df.columns
        if use_split:
            use_cols.append(split_col)

        temp = df[use_cols].copy()
        temp[variable] = to_numeric_series(temp[variable])
        temp = temp.dropna(subset=[x_col, variable])
        if temp.empty:
            continue

        if use_split:
            grouped = temp.groupby([x_col, split_col], dropna=False)[variable].mean().reset_index()
            grouped = grouped.rename(columns={variable: "Mean"})
            grouped["Split"] = grouped[split_col].fillna("Missing").astype(str)
        else:
            grouped = temp.groupby(x_col, dropna=False)[variable].mean().reset_index()
            grouped = grouped.rename(columns={variable: "Mean"})
            grouped["Split"] = variable

        grouped["Variable"] = variable
        rows.append(grouped[[x_col, "Variable", "Split", "Mean"]])

    if not rows:
        return pd.DataFrame(columns=[x_col, "Variable", "Split", "Mean"])

    result = pd.concat(rows, ignore_index=True)
    if sort_x:
        try:
            result = result.sort_values(["Variable", "Split", x_col]).reset_index(drop=True)
        except Exception:
            result["__x_sort__"] = result[x_col].astype(str)
            result = result.sort_values(["Variable", "Split", "__x_sort__"]).drop(columns="__x_sort__").reset_index(drop=True)
    return result


def pivot_for_variable(long_df: pd.DataFrame, x_col: str, variable: str) -> pd.DataFrame:
    temp = long_df[long_df["Variable"] == variable].copy()
    if temp.empty:
        return pd.DataFrame()
    pivot = temp.pivot_table(index=x_col, columns="Split", values="Mean", aggfunc="mean")
    try:
        pivot = pivot.sort_index()
    except Exception:
        ordered = sorted(pivot.index.tolist(), key=lambda x: str(x).lower())
        pivot = pivot.loc[ordered]
    return pivot


def apply_common_chart_style(ax, theme: Dict[str, str], show_grid: bool) -> None:
    ax.set_facecolor(theme["axes_facecolor"])
    ax.tick_params(axis="both", colors=theme["text_color"])
    ax.xaxis.label.set_color(theme["text_color"])
    ax.yaxis.label.set_color(theme["text_color"])
    ax.title.set_color(theme["text_color"])

    if show_grid:
        ax.grid(axis="both", linestyle="--", linewidth=0.6, alpha=0.45, color=theme["grid_color"])
        ax.set_axisbelow(True)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_color(theme["spine_color"])
    ax.spines["left"].set_color(theme["spine_color"])


def create_panel_mean_line_chart(
    long_df: pd.DataFrame,
    x_col: str,
    panel_variables: List[str],
    title: str,
    subtitle: str,
    x_label: str,
    y_label: str,
    figure_width: float,
    figure_height: float,
    panel_columns: int,
    share_y_axis: bool,
    theme_name: str,
    font_family: str,
    line_colors: Dict[str, str],
    line_width: float,
    show_markers: bool,
    marker_style: str,
    marker_size: int,
    show_value_labels: bool,
    value_label_decimal_digits: int,
    compact_labels: bool,
    title_font_size: int,
    subtitle_font_size: int,
    axis_font_size: int,
    tick_font_size: int,
    legend_font_size: int,
    value_label_font_size: int,
    panel_title_font_size: int,
    x_tick_rotation: int,
    show_grid: bool,
    legend_position: str,
) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)

    n_panels = len(panel_variables)
    ncols = max(1, min(panel_columns, n_panels))
    nrows = math.ceil(n_panels / ncols)

    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(figure_width, figure_height), sharey=share_y_axis, squeeze=False)
    fig.patch.set_facecolor(theme["figure_facecolor"])

    legend_handles = []
    legend_labels = []

    for panel_idx, variable in enumerate(panel_variables):
        row = panel_idx // ncols
        col = panel_idx % ncols
        ax = axes[row][col]

        pivot = pivot_for_variable(long_df, x_col=x_col, variable=variable)
        x_values = np.arange(len(pivot.index))
        x_labels = format_x_labels(pivot.index.tolist())

        for series_name in pivot.columns:
            y_values = pd.to_numeric(pivot[series_name], errors="coerce").to_numpy(dtype=float)
            color = line_colors.get(str(series_name), "#1F4E79")
            line, = ax.plot(
                x_values,
                y_values,
                label=str(series_name),
                color=color,
                linewidth=line_width,
                marker=marker_style if show_markers else None,
                markersize=marker_size if show_markers else 0,
                markeredgecolor="white",
                markeredgewidth=0.6,
            )
            if str(series_name) not in legend_labels:
                legend_handles.append(line)
                legend_labels.append(str(series_name))

            if show_value_labels:
                for x, y in zip(x_values, y_values):
                    if pd.isna(y) or not np.isfinite(y):
                        continue
                    ax.annotate(
                        format_numeric_label(y, value_label_decimal_digits, compact=compact_labels),
                        (x, y),
                        textcoords="offset points",
                        xytext=(0, 7),
                        ha="center",
                        fontsize=value_label_font_size,
                        color=theme["text_color"],
                    )

        ax.set_title(str(variable), fontsize=panel_title_font_size, fontweight="bold", color=theme["text_color"], pad=10)
        ax.set_xlabel(x_label, fontsize=axis_font_size, labelpad=8, color=theme["text_color"])
        ax.set_ylabel(y_label, fontsize=axis_font_size, labelpad=8, color=theme["text_color"])
        ax.set_xticks(x_values)
        ax.set_xticklabels(x_labels, rotation=x_tick_rotation, ha="right" if x_tick_rotation > 0 else "center", fontsize=tick_font_size, color=theme["text_color"])
        ax.tick_params(axis="y", labelsize=tick_font_size, colors=theme["text_color"])
        apply_common_chart_style(ax, theme, show_grid)

    for panel_idx in range(n_panels, nrows * ncols):
        row = panel_idx // ncols
        col = panel_idx % ncols
        axes[row][col].axis("off")

    fig.suptitle(title, fontsize=title_font_size, fontweight="bold", color=theme["text_color"], y=0.995)
    if subtitle.strip():
        fig.text(0.5, 0.965, subtitle, ha="center", va="top", fontsize=subtitle_font_size, color=theme["text_color"])

    if legend_handles:
        if legend_position == "Outside right":
            legend = fig.legend(legend_handles, legend_labels, loc="upper left", bbox_to_anchor=(1.005, 0.94), frameon=True, fontsize=legend_font_size)
            fig.tight_layout(rect=[0, 0, 0.86, 0.93])
        elif legend_position == "Bottom":
            legend = fig.legend(legend_handles, legend_labels, loc="lower center", bbox_to_anchor=(0.5, -0.005), ncol=min(4, len(legend_labels)), frameon=True, fontsize=legend_font_size)
            fig.tight_layout(rect=[0, 0.06, 1, 0.93])
        else:
            legend = fig.legend(legend_handles, legend_labels, loc="upper right", frameon=True, fontsize=legend_font_size)
            fig.tight_layout(rect=[0, 0, 1, 0.93])
        legend.get_frame().set_alpha(0.88)
        for text in legend.get_texts():
            text.set_color(theme["text_color"])
    else:
        fig.tight_layout(rect=[0, 0, 1, 0.93])

    return fig


# =====================================================
# LINEAR REGRESSION AND ASSUMPTION TESTS
# =====================================================

def significance_stars(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    if p_value < 0.10:
        return "†"
    return ""


def significance_decision(p_value: float, alpha: float) -> str:
    if pd.isna(p_value):
        return "Not available"
    return "Significant" if p_value < alpha else "Not significant"


def prepare_regression_data(df: pd.DataFrame, dependent_col: str, independent_cols: List[str]) -> Tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    reg_df = pd.DataFrame(index=df.index)
    reg_df[dependent_col] = to_numeric_series(df[dependent_col])
    for col in independent_cols:
        reg_df[col] = to_numeric_series(df[col])

    id_cols = [col for col in ["Ticker Code", "Company Name", "Year"] if col in df.columns]
    for col in id_cols:
        reg_df[col] = df[col]

    reg_df.insert(0, "Source Row", df.index)
    required_cols = [dependent_col] + independent_cols
    reg_df = reg_df.dropna(subset=required_cols).reset_index(drop=True)

    y = reg_df[dependent_col].astype(float)
    X = reg_df[independent_cols].astype(float)
    X = sm.add_constant(X, has_constant="add")
    return reg_df, y, X


def fit_ols_regression(df: pd.DataFrame, dependent_col: str, independent_cols: List[str]):
    reg_df, y, X = prepare_regression_data(df, dependent_col, independent_cols)
    model = sm.OLS(y, X).fit()
    return model, reg_df, y, X


def model_fit_summary_table(model, dependent_col: str, decimal_digits: int) -> pd.DataFrame:
    rmse = math.sqrt(float(np.mean(np.asarray(model.resid, dtype=float) ** 2))) if int(model.nobs) > 0 else np.nan
    rows = [{
        "Dependent Variable": dependent_col,
        "N": int(model.nobs),
        "R-Squared": float(getattr(model, "rsquared", np.nan)),
        "Adjusted R-Squared": float(getattr(model, "rsquared_adj", np.nan)),
        "F-statistic": float(getattr(model, "fvalue", np.nan)) if getattr(model, "fvalue", None) is not None else np.nan,
        "Prob(F-statistic)": float(getattr(model, "f_pvalue", np.nan)) if getattr(model, "f_pvalue", None) is not None else np.nan,
        "AIC": float(getattr(model, "aic", np.nan)),
        "BIC": float(getattr(model, "bic", np.nan)),
        "RMSE": rmse,
        "DF Model": float(getattr(model, "df_model", np.nan)),
        "DF Residual": float(getattr(model, "df_resid", np.nan)),
    }]
    result = pd.DataFrame(rows)
    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(decimal_digits)
    return result


def regression_coefficient_table(model, alpha: float, decimal_digits: int) -> pd.DataFrame:
    conf_int = model.conf_int(alpha=alpha)
    rows = []
    for var in model.params.index:
        p_value = float(model.pvalues[var]) if var in model.pvalues.index else np.nan
        rows.append({
            "Variable": var,
            "Coefficient": float(model.params[var]),
            "Std. Error": float(model.bse[var]),
            "t-statistic": float(model.tvalues[var]),
            "p-value": p_value,
            "Significance": significance_stars(p_value),
            "Decision": significance_decision(p_value, alpha),
            "CI Lower": float(conf_int.loc[var, 0]) if var in conf_int.index else np.nan,
            "CI Upper": float(conf_int.loc[var, 1]) if var in conf_int.index else np.nan,
        })
    result = pd.DataFrame(rows)
    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(decimal_digits)
    return result


def robust_regression_coefficient_table(model, cov_type: str, alpha: float, decimal_digits: int) -> pd.DataFrame:
    robust_model = model.get_robustcov_results(cov_type=cov_type)
    params = pd.Series(robust_model.params, index=model.params.index)
    bse = pd.Series(robust_model.bse, index=model.params.index)
    tvalues = pd.Series(robust_model.tvalues, index=model.params.index)
    pvalues = pd.Series(robust_model.pvalues, index=model.params.index)
    ci = pd.DataFrame(robust_model.conf_int(alpha=alpha), index=model.params.index, columns=["CI Lower", "CI Upper"])
    rows = []
    for var in model.params.index:
        p_value = float(pvalues[var]) if var in pvalues.index else np.nan
        rows.append({
            "Variable": var,
            "Coefficient": float(params[var]),
            "Robust Std. Error": float(bse[var]),
            "t-statistic": float(tvalues[var]),
            "p-value": p_value,
            "Significance": significance_stars(p_value),
            "Decision": significance_decision(p_value, alpha),
            "CI Lower": float(ci.loc[var, "CI Lower"]),
            "CI Upper": float(ci.loc[var, "CI Upper"]),
        })
    result = pd.DataFrame(rows)
    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(decimal_digits)
    return result


def normality_tests(residuals: np.ndarray, alpha: float, decimal_digits: int) -> pd.DataFrame:
    residuals = np.asarray(residuals, dtype=float)
    residuals = residuals[np.isfinite(residuals)]
    rows = []
    n = len(residuals)
    if n < 3:
        return pd.DataFrame(columns=["Test", "Statistic", "p-value", "Decision", "Note"])

    std = np.std(residuals, ddof=1)
    z = (residuals - np.mean(residuals)) / std if std > 0 else residuals

    for method_label, method in [("Kolmogorov-Smirnov Exact", "exact"), ("Kolmogorov-Smirnov Asymptotic", "asymp")]:
        try:
            stat, p = stats.kstest(z, "norm", method=method)
            rows.append({
                "Test": method_label,
                "Statistic": stat,
                "p-value": p,
                "Decision": "Normal residuals" if p >= alpha else "Non-normal residuals",
                "Note": "Based on standardized residuals.",
            })
        except Exception as exc:
            rows.append({"Test": method_label, "Statistic": np.nan, "p-value": np.nan, "Decision": "Not available", "Note": str(exc)})

    try:
        jb = stats.jarque_bera(residuals)
        rows.append({
            "Test": "Jarque-Bera",
            "Statistic": float(jb.statistic),
            "p-value": float(jb.pvalue),
            "Decision": "Normal residuals" if jb.pvalue >= alpha else "Non-normal residuals",
            "Note": "Skewness-kurtosis normality test.",
        })
    except Exception as exc:
        rows.append({"Test": "Jarque-Bera", "Statistic": np.nan, "p-value": np.nan, "Decision": "Not available", "Note": str(exc)})

    try:
        if n <= 5000:
            shapiro_stat, shapiro_p = stats.shapiro(residuals)
            note = ""
        else:
            sample = residuals[:5000]
            shapiro_stat, shapiro_p = stats.shapiro(sample)
            note = "Computed on the first 5000 residuals due to sample-size guidance."
        rows.append({
            "Test": "Shapiro-Wilk",
            "Statistic": float(shapiro_stat),
            "p-value": float(shapiro_p),
            "Decision": "Normal residuals" if shapiro_p >= alpha else "Non-normal residuals",
            "Note": note,
        })
    except Exception as exc:
        rows.append({"Test": "Shapiro-Wilk", "Statistic": np.nan, "p-value": np.nan, "Decision": "Not available", "Note": str(exc)})

    try:
        ad = stats.anderson(residuals, dist="norm")
        sig_levels = list(ad.significance_level)
        cv_5 = ad.critical_values[sig_levels.index(5.0)] if 5.0 in sig_levels else np.nan
        rows.append({
            "Test": "Anderson-Darling",
            "Statistic": float(ad.statistic),
            "p-value": np.nan,
            "Decision": "Normal residuals" if np.isfinite(cv_5) and ad.statistic < cv_5 else "Non-normal residuals",
            "Note": f"Critical value at 5% = {cv_5:.{decimal_digits}f}; SciPy does not return a direct p-value.",
        })
    except Exception as exc:
        rows.append({"Test": "Anderson-Darling", "Statistic": np.nan, "p-value": np.nan, "Decision": "Not available", "Note": str(exc)})

    result = pd.DataFrame(rows)
    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(decimal_digits)
    return result


def multicollinearity_diagnostics(reg_df: pd.DataFrame, independent_cols: List[str], decimal_digits: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    X = reg_df[independent_cols].apply(pd.to_numeric, errors="coerce").dropna()
    X_const = sm.add_constant(X, has_constant="add")
    rows = []
    for i, col in enumerate(X_const.columns):
        if col == "const":
            continue
        try:
            vif = float(variance_inflation_factor(X_const.values, i))
        except Exception:
            vif = np.nan
        rows.append({
            "Variable": col,
            "VIF": vif,
            "Tolerance": 1 / vif if pd.notna(vif) and vif != 0 else np.nan,
            "Common Rule": "Potential multicollinearity" if pd.notna(vif) and vif > 10 else "Acceptable",
        })
    vif_df = pd.DataFrame(rows)
    numeric_cols = vif_df.select_dtypes(include=[np.number]).columns
    vif_df[numeric_cols] = vif_df[numeric_cols].round(decimal_digits)
    corr_df = X.corr().round(decimal_digits)
    return vif_df, corr_df


def runs_test(residuals: np.ndarray, alpha: float) -> pd.DataFrame:
    residuals = np.asarray(residuals, dtype=float)
    residuals = residuals[np.isfinite(residuals)]
    median = np.median(residuals)
    signs = residuals > median
    signs = signs[residuals != median]
    n1 = int(np.sum(signs))
    n2 = int(len(signs) - n1)

    if n1 == 0 or n2 == 0 or len(signs) < 2:
        runs = np.nan
        z_stat = np.nan
        p_value = np.nan
    else:
        runs = 1 + int(np.sum(signs[1:] != signs[:-1]))
        expected_runs = 1 + (2 * n1 * n2) / (n1 + n2)
        variance_runs = (2 * n1 * n2 * (2 * n1 * n2 - n1 - n2)) / (((n1 + n2) ** 2) * (n1 + n2 - 1))
        z_stat = (runs - expected_runs) / math.sqrt(variance_runs) if variance_runs > 0 else np.nan
        p_value = 2 * stats.norm.sf(abs(z_stat)) if np.isfinite(z_stat) else np.nan

    return pd.DataFrame([{
        "Test": "Runs Test",
        "Statistic": z_stat,
        "Runs": runs,
        "p-value": p_value,
        "Decision": "No autocorrelation pattern detected" if pd.notna(p_value) and p_value >= alpha else "Autocorrelation pattern detected",
    }])


def autocorrelation_tests(residuals: np.ndarray, alpha: float, decimal_digits: int) -> pd.DataFrame:
    runs_df = runs_test(residuals, alpha)
    try:
        dw_value = durbin_watson(residuals)
    except Exception:
        dw_value = np.nan
    dw_df = pd.DataFrame([{
        "Test": "Durbin-Watson",
        "Statistic": float(dw_value) if pd.notna(dw_value) else np.nan,
        "Runs": np.nan,
        "p-value": np.nan,
        "Decision": "Approximately no autocorrelation if close to 2; values near 0 indicate positive autocorrelation; values near 4 indicate negative autocorrelation.",
    }])
    result = pd.concat([runs_df, dw_df], ignore_index=True)
    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(decimal_digits)
    return result


def glejser_test(residuals: np.ndarray, X: pd.DataFrame, alpha: float, decimal_digits: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    abs_resid = np.abs(np.asarray(residuals, dtype=float))
    model = sm.OLS(abs_resid, X).fit()
    rows = []
    for var in model.params.index:
        p_value = float(model.pvalues[var]) if var in model.pvalues.index else np.nan
        rows.append({
            "Variable": var,
            "Coefficient": float(model.params[var]),
            "t-statistic": float(model.tvalues[var]),
            "p-value": p_value,
            "Decision": significance_decision(p_value, alpha),
        })
    coef_df = pd.DataFrame(rows)
    summary_df = pd.DataFrame([{
        "Test": "Glejser Test",
        "Statistic": float(model.fvalue) if model.fvalue is not None else np.nan,
        "p-value": float(model.f_pvalue) if model.f_pvalue is not None else np.nan,
        "Decision": "Heteroskedasticity detected" if model.f_pvalue is not None and model.f_pvalue < alpha else "No heteroskedasticity detected",
    }])
    for df_out in [coef_df, summary_df]:
        numeric_cols = df_out.select_dtypes(include=[np.number]).columns
        df_out[numeric_cols] = df_out[numeric_cols].round(decimal_digits)
    return coef_df, summary_df


def heteroskedasticity_tests(residuals: np.ndarray, X: pd.DataFrame, alpha: float, decimal_digits: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    try:
        bp_lm, bp_lm_p, bp_f, bp_f_p = het_breuschpagan(residuals, X)
        rows.append({
            "Test": "Breusch-Pagan",
            "LM Statistic": bp_lm,
            "LM p-value": bp_lm_p,
            "F Statistic": bp_f,
            "F p-value": bp_f_p,
            "Decision": "Heteroskedasticity detected" if bp_lm_p < alpha else "No heteroskedasticity detected",
        })
    except Exception as exc:
        rows.append({"Test": "Breusch-Pagan", "LM Statistic": np.nan, "LM p-value": np.nan, "F Statistic": np.nan, "F p-value": np.nan, "Decision": str(exc)})

    try:
        white_lm, white_lm_p, white_f, white_f_p = het_white(residuals, X)
        rows.append({
            "Test": "White Test",
            "LM Statistic": white_lm,
            "LM p-value": white_lm_p,
            "F Statistic": white_f,
            "F p-value": white_f_p,
            "Decision": "Heteroskedasticity detected" if white_lm_p < alpha else "No heteroskedasticity detected",
        })
    except Exception as exc:
        rows.append({"Test": "White Test", "LM Statistic": np.nan, "LM p-value": np.nan, "F Statistic": np.nan, "F p-value": np.nan, "Decision": str(exc)})

    glejser_coef, glejser_summary = glejser_test(residuals, X, alpha, decimal_digits)
    glejser_row = {
        "Test": "Glejser Test",
        "LM Statistic": glejser_summary.loc[0, "Statistic"] if not glejser_summary.empty else np.nan,
        "LM p-value": glejser_summary.loc[0, "p-value"] if not glejser_summary.empty else np.nan,
        "F Statistic": np.nan,
        "F p-value": np.nan,
        "Decision": glejser_summary.loc[0, "Decision"] if not glejser_summary.empty else "Not available",
    }
    rows.append(glejser_row)
    summary = pd.DataFrame(rows)
    numeric_cols = summary.select_dtypes(include=[np.number]).columns
    summary[numeric_cols] = summary[numeric_cols].round(decimal_digits)
    return summary, glejser_coef


def outlier_influence_table(model, reg_df: pd.DataFrame, alpha: float, decimal_digits: int) -> pd.DataFrame:
    try:
        influence = OLSInfluence(model)
        studentized = influence.resid_studentized_external
        leverage = influence.hat_matrix_diag
        cooks = influence.cooks_distance[0]
    except Exception:
        residuals = np.asarray(model.resid, dtype=float)
        studentized = (residuals - np.mean(residuals)) / np.std(residuals, ddof=1)
        leverage = np.repeat(np.nan, len(residuals))
        cooks = np.repeat(np.nan, len(residuals))

    n = len(reg_df)
    p = int(getattr(model, "df_model", 1)) + 1
    leverage_cutoff = 2 * p / max(n, 1)
    cooks_cutoff = 4 / max(n, 1)

    result = pd.DataFrame({
        "Source Row": reg_df["Source Row"].values if "Source Row" in reg_df.columns else np.arange(n),
        "Residual": np.asarray(model.resid, dtype=float),
        "Fitted Value": np.asarray(model.fittedvalues, dtype=float),
        "Studentized Residual": np.asarray(studentized, dtype=float),
        "Leverage": np.asarray(leverage, dtype=float),
        "Cook's Distance": np.asarray(cooks, dtype=float),
    })
    for col in ["Ticker Code", "Company Name", "Year"]:
        if col in reg_df.columns:
            result.insert(1, col, reg_df[col].values)
    result["Residual Outlier"] = np.abs(result["Studentized Residual"]) > 3
    result["High Leverage"] = result["Leverage"] > leverage_cutoff
    result["Influential by Cook"] = result["Cook's Distance"] > cooks_cutoff
    result["Any Flag"] = result[["Residual Outlier", "High Leverage", "Influential by Cook"]].any(axis=1)
    result["Rule"] = f"|Studentized residual| > 3; leverage > {leverage_cutoff:.4f}; Cook's D > {cooks_cutoff:.4f}"
    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(decimal_digits)
    return result


def create_regression_diagnostic_figure(model, theme_name: str, font_family: str, figure_width: float, figure_height: float, show_grid: bool) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)
    fitted = np.asarray(model.fittedvalues, dtype=float)
    residuals = np.asarray(model.resid, dtype=float)
    std_resid = (residuals - np.mean(residuals)) / np.std(residuals, ddof=1) if np.std(residuals, ddof=1) > 0 else residuals

    fig, axes = plt.subplots(1, 2, figsize=(figure_width, figure_height), squeeze=False)
    fig.patch.set_facecolor(theme["figure_facecolor"])

    ax1 = axes[0][0]
    ax1.scatter(fitted, residuals, s=55, alpha=0.8, edgecolor="white", linewidth=0.5)
    ax1.axhline(0, color=theme["spine_color"], linestyle="--", linewidth=1.0)
    ax1.set_title("Residuals vs Fitted Values", fontsize=13, fontweight="bold", color=theme["text_color"])
    ax1.set_xlabel("Fitted value", color=theme["text_color"])
    ax1.set_ylabel("Residual", color=theme["text_color"])
    apply_common_chart_style(ax1, theme, show_grid)

    ax2 = axes[0][1]
    stats.probplot(std_resid[np.isfinite(std_resid)], dist="norm", plot=ax2)
    ax2.set_title("Normal Q-Q Plot", fontsize=13, fontweight="bold", color=theme["text_color"])
    ax2.get_lines()[0].set_markerfacecolor("#1F4E79")
    ax2.get_lines()[0].set_markeredgecolor("white")
    ax2.get_lines()[0].set_alpha(0.8)
    ax2.get_lines()[1].set_color(theme["spine_color"])
    ax2.tick_params(axis="both", colors=theme["text_color"])
    ax2.xaxis.label.set_color(theme["text_color"])
    ax2.yaxis.label.set_color(theme["text_color"])
    apply_common_chart_style(ax2, theme, show_grid)

    fig.suptitle("Regression Residual Diagnostic Plots", fontsize=16, fontweight="bold", color=theme["text_color"])
    fig.tight_layout(rect=[0, 0, 1, 0.92])
    return fig

# =====================================================
# EXPORT HELPERS
# =====================================================

def fig_to_png_bytes(fig, dpi: int, transparent_background: bool = False) -> bytes:
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="none" if transparent_background else fig.get_facecolor(),
        transparent=transparent_background,
    )
    buffer.seek(0)
    return buffer.getvalue()


def style_excel_workbook(workbook) -> None:
    header_fill = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 55)


def dataframe_to_excel_bytes(sheets: Dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            safe_sheet = re.sub(r"[^A-Za-z0-9 _-]", "", str(sheet_name)).strip()[:31]
            if not safe_sheet:
                safe_sheet = "Sheet"
            data.to_excel(writer, sheet_name=safe_sheet, index=False)
        style_excel_workbook(writer.book)
    output.seek(0)
    return output.getvalue()


# =====================================================
# HEADER
# =====================================================

with st.container():
    left, right = st.columns([1, 5])
    with left:
        if LOGO_PATH.exists():
            safe_streamlit_image(LOGO_PATH, width=210)
        else:
            st.markdown("### STATCAL")
    with right:
        st.title(APP_NAME)
        st.subheader(APP_TITLE)
        st.caption(
            "A Python Streamlit application for exploring IDX banking financial data through data filtering, "
            "descriptive statistics, grouped summaries, publication-ready multi-panel line charts, "
            "linear regression with assumption testing, and exportable analytical outputs."
        )

st.markdown(
    f"""
    **Website:** [{WEBSITE_URL}]({WEBSITE_URL})  
    **STATCAL ONLINE Page:** [{STATCAL_ONLINE_URL}]({STATCAL_ONLINE_URL})  
    **Data Source / Training Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})  
    **{APP_UPDATED}**  
    **Purpose:** Explore IDX banking financial data by year and ticker code, create descriptive tables, multi-panel line charts, regression analysis, assumption tests, and export outputs for academic reporting.

    ---
    """
)


# =====================================================
# TABS
# =====================================================

tab_data, tab_uni, tab_group, tab_line, tab_regression, tab_export = st.tabs([
    "1. Data & Filters",
    "2. Univariate Descriptive",
    "3. Grouped Descriptive",
    "4. Multi-Panel Line Chart",
    "5. Regression & Assumption Tests",
    "6. Export Charts & Excel",
])


# =====================================================
# TAB 1: DATA INPUT AND FILTERS
# =====================================================

with tab_data:
    st.subheader("Data Input and Flexible Filters")
    st.write("Upload an Excel dataset or use the included sample dataset: **data idx perbankan.xlsx**.")
    st.markdown(f"**Training Data / Sample Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})")

    uploaded_file = st.file_uploader(
        "Upload Excel file",
        type=["xlsx", "xls"],
        help="Upload financial data in Excel format.",
    )

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        source_name = uploaded_file.name
    elif SAMPLE_DATA_PATH.exists():
        file_bytes = SAMPLE_DATA_PATH.read_bytes()
        source_name = SAMPLE_DATA_PATH.name
    else:
        st.info("Please upload an Excel file to start the analysis.")
        st.stop()

    try:
        sheet_names = get_excel_sheet_names(file_bytes)
        sheet_name = st.selectbox("Worksheet", sheet_names, index=0)
        raw_df = read_excel_file(file_bytes, sheet_name)
        df = clean_dataframe(raw_df)
    except Exception as exc:
        st.error("Failed to read the Excel file.")
        st.exception(exc)
        st.stop()

    if df.empty:
        st.error("The selected worksheet is empty.")
        st.stop()

    columns = df.columns.tolist()
    numeric_candidates = detect_numeric_columns(df)

    st.markdown("#### Filter Settings")
    filter_col_1, filter_col_2 = st.columns(2)

    if "Year" in columns:
        year_values = sorted_unique_values(df["Year"])
        with filter_col_1:
            selected_years = st.multiselect("Filter by Year", year_values, default=year_values, format_func=lambda x: str(x))
    else:
        selected_years = []

    if "Ticker Code" in columns:
        ticker_values = sorted_unique_values(df["Ticker Code"])
        with filter_col_2:
            selected_tickers = st.multiselect("Filter by Ticker Code", ticker_values, default=ticker_values, format_func=lambda x: str(x))
    else:
        selected_tickers = []

    filtered_df = df.copy()
    if "Year" in columns and selected_years:
        filtered_df = filtered_df[filtered_df["Year"].isin(selected_years)]
    if "Ticker Code" in columns and selected_tickers:
        filtered_df = filtered_df[filtered_df["Ticker Code"].isin(selected_tickers)]

    if filtered_df.empty:
        st.error("No rows remain after filtering. Please adjust the Year or Ticker Code filter.")
        st.stop()

    st.markdown("#### Dataset Preview")
    st.write(f"**Source:** {source_name} | **Worksheet:** {sheet_name}")
    st.dataframe(make_arrow_safe_dataframe(filtered_df), use_container_width=True)

    metric_1, metric_2, metric_3, metric_4 = st.columns(4)
    with metric_1:
        st.metric("Original rows", f"{len(df):,}")
    with metric_2:
        st.metric("Rows after filtering", f"{len(filtered_df):,}")
    with metric_3:
        st.metric("Columns", f"{len(df.columns):,}")
    with metric_4:
        st.metric("Detected numeric variables", f"{len(numeric_candidates):,}")

    if FINANCIAL_URL_COL in filtered_df.columns:
        with st.expander("Financial Statement URL List", expanded=False):
            url_df = filtered_df[["Ticker Code", "Company Name", "Year", FINANCIAL_URL_COL]].copy() if all(c in filtered_df.columns for c in ["Ticker Code", "Company Name", "Year"]) else filtered_df[[FINANCIAL_URL_COL]].copy()
            st.dataframe(make_arrow_safe_dataframe(url_df), use_container_width=True)

    with st.expander("Detected numeric variables", expanded=False):
        st.write(numeric_candidates)


# =====================================================
# TAB 2: UNIVARIATE DESCRIPTIVE
# =====================================================

with tab_uni:
    st.subheader("Univariate Descriptive Statistics")

    if not numeric_candidates:
        st.warning("No numeric variables were detected.")
        st.stop()

    uni_col_1, uni_col_2 = st.columns([3, 1])
    with uni_col_1:
        uni_numeric_cols = st.multiselect(
            "Select numeric variables for univariate descriptive statistics",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="uni_numeric_cols",
        )
    with uni_col_2:
        uni_decimal_digits = st.slider("Decimal digits", 0, 8, 3, 1, key="uni_decimal_digits")

    if not uni_numeric_cols:
        st.warning("Please select at least one numeric variable.")
    else:
        descriptive_stats = compute_descriptive_statistics(filtered_df, uni_numeric_cols, uni_decimal_digits)
        st.dataframe(make_arrow_safe_dataframe(descriptive_stats), use_container_width=True)


# =====================================================
# TAB 3: GROUPED DESCRIPTIVE
# =====================================================

with tab_group:
    st.subheader("Grouped Descriptive Statistics")

    group_col_1, group_col_2 = st.columns([3, 2])
    with group_col_1:
        group_numeric_cols = st.multiselect(
            "Select numeric variables for grouped descriptive statistics",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="group_numeric_cols",
        )
    with group_col_2:
        default_group = [c for c in ["Year", "Ticker Code"] if c in columns]
        group_cols = st.multiselect(
            "Group by category variables",
            columns,
            default=default_group,
            key="group_cols",
        )

    group_decimal_digits = st.slider("Decimal digits for grouped statistics", 0, 8, 3, 1, key="group_decimal_digits")

    if not group_numeric_cols:
        st.warning("Please select at least one numeric variable.")
    elif not group_cols:
        st.info("Please select at least one group variable to generate grouped descriptive statistics.")
        grouped_descriptive_stats = pd.DataFrame()
    else:
        grouped_descriptive_stats = compute_grouped_descriptive_statistics(filtered_df, group_numeric_cols, group_cols, group_decimal_digits)
        st.dataframe(make_arrow_safe_dataframe(grouped_descriptive_stats), use_container_width=True)


# =====================================================
# TAB 4: MULTI-PANEL LINE CHART
# =====================================================

with tab_line:
    st.subheader("Multi-Panel Mean Line Chart")

    line_setting_1, line_setting_2, line_setting_3 = st.columns(3)
    with line_setting_1:
        x_index = preferred_option(columns, ["Year", "Tahun", "Date", "Time", "Period", "Periode"], fallback_index=0)
        line_x_col = st.selectbox("X-axis variable", columns, index=x_index, key="line_x_col")
        line_numeric_cols = st.multiselect(
            "Panel variables / numeric variables",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates)[:4],
            key="line_numeric_cols",
        )
    with line_setting_2:
        split_options = ["None"] + columns
        split_index = preferred_option(split_options, ["Ticker Code", "Company Name", "Year"], fallback_index=0)
        line_split_col = st.selectbox("Split lines by category", split_options, index=split_index, key="line_split_col")
        line_sort_x = st.checkbox("Sort X-axis", value=True, key="line_sort_x")
    with line_setting_3:
        line_max_panels = st.slider("Maximum panels", 1, 20, min(6, max(1, len(line_numeric_cols))), key="line_max_panels")
        line_panel_columns = st.slider("Number of panel columns", 1, 4, min(2, max(1, line_max_panels)), key="line_panel_cols")
        line_share_y_axis = st.checkbox("Share Y-axis across panels", value=False, key="line_share_y")

    style_expander = st.expander("Line Chart Settings", expanded=True)
    with style_expander:
        style_col_1, style_col_2, style_col_3 = st.columns(3)
        with style_col_1:
            line_font_family = st.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2, key="line_font")
            line_theme = st.selectbox("Chart background theme", list(THEMES.keys()), index=0, key="line_theme")
            line_palette = st.selectbox("Color palette", list(COLOR_PALETTES.keys()), index=0, key="line_palette")
        with style_col_2:
            line_fig_width = st.slider("Figure width", 6.0, 30.0, 15.0, 0.5, key="line_fig_width")
            line_fig_height = st.slider("Figure height", 4.0, 30.0, 10.0, 0.5, key="line_fig_height")
            line_width = st.slider("Line width", 0.5, 8.0, 2.4, 0.1, key="line_width")
        with style_col_3:
            line_show_markers = st.checkbox("Add markers to lines", value=True, key="line_markers")
            line_marker_style = st.selectbox("Marker style", MARKERS, index=0, key="line_marker_style")
            line_marker_size = st.slider("Marker size", 2, 18, 6, key="line_marker_size")

        line_title = st.text_area("Chart title", value="Mean Trend of IDX Banking Financial Variables", height=68, key="line_title")
        line_subtitle = st.text_input("Chart subtitle", value="Multi-panel chart based on mean values from filtered IDX banking data", key="line_subtitle")
        line_x_label = st.text_input("X-axis label", value=line_x_col, key="line_x_label")
        line_y_label = st.text_input("Y-axis label", value="Mean", key="line_y_label")

        line_show_value_labels = st.checkbox("Show mean values on chart", value=False, key="line_show_value")
        line_compact_labels = st.checkbox("Use compact value labels (K/M/B/T)", value=True, key="line_compact")
        line_x_tick_rotation = st.slider("X-axis label rotation", 0, 90, 0, 5, key="line_x_rot")
        line_show_grid = st.checkbox("Show grid", value=True, key="line_grid")
        line_legend_position = st.selectbox("Legend position", ["Outside right", "Bottom", "Best"], index=0, key="line_legend")

        line_title_font_size = st.slider("Main title font size", 10, 44, 18, key="line_title_font")
        line_subtitle_font_size = st.slider("Subtitle font size", 8, 30, 11, key="line_subtitle_font")
        line_panel_title_font_size = st.slider("Panel title font size", 8, 30, 13, key="line_panel_title_font")
        line_axis_font_size = st.slider("Axis label font size", 8, 30, 12, key="line_axis_font")
        line_tick_font_size = st.slider("Tick label font size", 6, 26, 10, key="line_tick_font")
        line_legend_font_size = st.slider("Legend font size", 6, 24, 9, key="line_legend_font")
        line_value_label_font_size = st.slider("Value label font size", 5, 22, 8, key="line_value_font")

    line_panel_variables = line_numeric_cols[:line_max_panels] if line_numeric_cols else []
    if not line_panel_variables:
        st.warning("Please select at least one numeric variable for the line chart.")
        chart_long_data = pd.DataFrame()
    else:
        chart_long_data = build_mean_line_long_data(
            filtered_df,
            x_col=line_x_col,
            numeric_cols=line_panel_variables,
            split_col=line_split_col,
            sort_x=line_sort_x,
        )

        if chart_long_data.empty:
            st.error("The mean line chart data is empty. Please adjust variables or filters.")
        else:
            line_series = sorted(chart_long_data["Split"].dropna().astype(str).unique().tolist(), key=lambda x: str(x).lower())
            palette = get_palette_color_list(line_palette)
            line_colors: Dict[str, str] = {}
            st.markdown("#### Line Colors")
            color_cols = st.columns(4)
            for idx, series_name in enumerate(line_series[:32]):
                with color_cols[idx % 4]:
                    line_colors[str(series_name)] = st.color_picker(f"{series_name}", palette[idx % len(palette)], key=f"line_color_{idx}_{series_name}")

            with st.expander("Mean line chart data preview", expanded=False):
                st.dataframe(make_arrow_safe_dataframe(chart_long_data.head(300)), use_container_width=True)

            line_fig = create_panel_mean_line_chart(
                long_df=chart_long_data,
                x_col=line_x_col,
                panel_variables=line_panel_variables,
                title=line_title,
                subtitle=line_subtitle,
                x_label=line_x_label,
                y_label=line_y_label,
                figure_width=line_fig_width,
                figure_height=line_fig_height,
                panel_columns=line_panel_columns,
                share_y_axis=line_share_y_axis,
                theme_name=line_theme,
                font_family=line_font_family,
                line_colors=line_colors,
                line_width=line_width,
                show_markers=line_show_markers,
                marker_style=line_marker_style,
                marker_size=line_marker_size,
                show_value_labels=line_show_value_labels,
                value_label_decimal_digits=3,
                compact_labels=line_compact_labels,
                title_font_size=line_title_font_size,
                subtitle_font_size=line_subtitle_font_size,
                axis_font_size=line_axis_font_size,
                tick_font_size=line_tick_font_size,
                legend_font_size=line_legend_font_size,
                value_label_font_size=line_value_label_font_size,
                panel_title_font_size=line_panel_title_font_size,
                x_tick_rotation=line_x_tick_rotation,
                show_grid=line_show_grid,
                legend_position=line_legend_position,
            )
            st.pyplot(line_fig)


# =====================================================
# TAB 5: REGRESSION AND ASSUMPTION TESTS
# =====================================================

with tab_regression:
    st.subheader("Linear Regression and Classical Assumption Tests")
    st.write(
        "This tab estimates a multiple linear regression model and provides diagnostic tests for normality, "
        "multicollinearity, heteroskedasticity, autocorrelation, and influential observations."
    )

    if len(numeric_candidates) < 2:
        st.warning("At least two numeric variables are required for regression analysis.")
        regression_summary_table = pd.DataFrame()
        regression_coef_table = pd.DataFrame()
        robust_coef_table = pd.DataFrame()
        normality_df = pd.DataFrame()
        vif_df = pd.DataFrame()
        independent_corr_df = pd.DataFrame()
        hetero_summary_df = pd.DataFrame()
        glejser_coef_df = pd.DataFrame()
        autocorr_df = pd.DataFrame()
        outlier_df = pd.DataFrame()
    else:
        reg_col_1, reg_col_2, reg_col_3 = st.columns([2, 3, 1])
        with reg_col_1:
            dep_index = preferred_option(numeric_candidates, ["Net Income", "Return on Asset (ROA)", "Total Assets"], fallback_index=0)
            regression_dependent = st.selectbox("Dependent variable", numeric_candidates, index=dep_index, key="reg_dep")
        with reg_col_2:
            reg_ind_options = [col for col in numeric_candidates if col != regression_dependent]
            preferred_reg_ind = [col for col in reg_ind_options if col in ["Total Assets", "Total Liabilities", "Total Equity", "Earning per Share", "Debt to Asset Ratio (DAR)", "Debt to Equity Ratio (DER)"]]
            regression_independent = st.multiselect(
                "Independent variables",
                reg_ind_options,
                default=preferred_reg_ind[:3] if preferred_reg_ind else reg_ind_options[:3],
                key="reg_ind",
            )
        with reg_col_3:
            reg_decimal_digits = st.slider("Decimal digits", 0, 8, 4, 1, key="reg_digits")
            reg_alpha = st.selectbox("Alpha", [0.10, 0.05, 0.01], index=1, key="reg_alpha")

        reg_settings = st.expander("Regression Settings", expanded=True)
        with reg_settings:
            reg_setting_1, reg_setting_2, reg_setting_3 = st.columns(3)
            with reg_setting_1:
                robust_cov_type = st.selectbox("Robust standard error", ["HC0", "HC1", "HC2", "HC3"], index=1, key="reg_robust_cov")
            with reg_setting_2:
                reg_theme = st.selectbox("Residual plot theme", list(THEMES.keys()), index=0, key="reg_theme")
                reg_font_family = st.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2, key="reg_font")
            with reg_setting_3:
                reg_fig_width = st.slider("Residual plot width", 7.0, 22.0, 14.0, 0.5, key="reg_fig_width")
                reg_fig_height = st.slider("Residual plot height", 3.5, 12.0, 5.5, 0.5, key="reg_fig_height")
                reg_show_grid = st.checkbox("Show grid", value=True, key="reg_grid")

        if not regression_independent:
            st.warning("Please select at least one independent variable.")
            regression_summary_table = pd.DataFrame()
            regression_coef_table = pd.DataFrame()
            robust_coef_table = pd.DataFrame()
            normality_df = pd.DataFrame()
            vif_df = pd.DataFrame()
            independent_corr_df = pd.DataFrame()
            hetero_summary_df = pd.DataFrame()
            glejser_coef_df = pd.DataFrame()
            autocorr_df = pd.DataFrame()
            outlier_df = pd.DataFrame()
        else:
            try:
                regression_model, regression_model_data, regression_y, regression_X = fit_ols_regression(filtered_df, regression_dependent, regression_independent)

                regression_summary_table = model_fit_summary_table(regression_model, regression_dependent, reg_decimal_digits)
                regression_coef_table = regression_coefficient_table(regression_model, reg_alpha, reg_decimal_digits)
                robust_coef_table = robust_regression_coefficient_table(regression_model, robust_cov_type, reg_alpha, reg_decimal_digits)
                normality_df = normality_tests(regression_model.resid, reg_alpha, reg_decimal_digits)
                vif_df, independent_corr_df = multicollinearity_diagnostics(regression_model_data, regression_independent, reg_decimal_digits)
                hetero_summary_df, glejser_coef_df = heteroskedasticity_tests(np.asarray(regression_model.resid, dtype=float), regression_X, reg_alpha, reg_decimal_digits)
                autocorr_df = autocorrelation_tests(np.asarray(regression_model.resid, dtype=float), reg_alpha, reg_decimal_digits)
                outlier_df = outlier_influence_table(regression_model, regression_model_data, reg_alpha, reg_decimal_digits)
                flagged_outlier_df = outlier_df[outlier_df["Any Flag"]].copy() if not outlier_df.empty else pd.DataFrame()

                st.markdown("#### Model Fit Summary")
                st.dataframe(make_arrow_safe_dataframe(regression_summary_table), use_container_width=True)

                st.markdown("#### Coefficient Estimates: Classical Standard Errors")
                st.dataframe(make_arrow_safe_dataframe(regression_coef_table), use_container_width=True)

                st.markdown(f"#### Coefficient Estimates: {robust_cov_type} Robust Standard Errors")
                st.dataframe(make_arrow_safe_dataframe(robust_coef_table), use_container_width=True)

                diag_1, diag_2 = st.columns(2)
                with diag_1:
                    st.markdown("#### Normality Tests")
                    st.dataframe(make_arrow_safe_dataframe(normality_df), use_container_width=True)

                    st.markdown("#### Multicollinearity: VIF and Tolerance")
                    st.dataframe(make_arrow_safe_dataframe(vif_df), use_container_width=True)

                    st.markdown("#### Autocorrelation Tests")
                    st.dataframe(make_arrow_safe_dataframe(autocorr_df), use_container_width=True)

                with diag_2:
                    st.markdown("#### Heteroskedasticity Tests")
                    st.dataframe(make_arrow_safe_dataframe(hetero_summary_df), use_container_width=True)

                    st.markdown("#### Glejser Test Coefficients")
                    st.dataframe(make_arrow_safe_dataframe(glejser_coef_df), use_container_width=True)

                    st.markdown("#### Correlation among Independent Variables")
                    st.dataframe(make_arrow_safe_dataframe(independent_corr_df.reset_index().rename(columns={"index": "Variable"})), use_container_width=True)

                st.markdown("#### Outlier and Influential Observation Detection")
                out_col_1, out_col_2, out_col_3 = st.columns(3)
                with out_col_1:
                    st.metric("Residual outliers", f"{int(outlier_df['Residual Outlier'].sum()) if not outlier_df.empty else 0:,}")
                with out_col_2:
                    st.metric("High leverage", f"{int(outlier_df['High Leverage'].sum()) if not outlier_df.empty else 0:,}")
                with out_col_3:
                    st.metric("Influential by Cook", f"{int(outlier_df['Influential by Cook'].sum()) if not outlier_df.empty else 0:,}")

                if flagged_outlier_df.empty:
                    st.success("No flagged observations were detected using the selected diagnostic rules.")
                else:
                    st.dataframe(make_arrow_safe_dataframe(flagged_outlier_df), use_container_width=True)

                with st.expander("Full residual and influence diagnostic table", expanded=False):
                    st.dataframe(make_arrow_safe_dataframe(outlier_df), use_container_width=True)

                st.markdown("#### Residual Diagnostic Plots")
                regression_residual_fig = create_regression_diagnostic_figure(
                    regression_model,
                    theme_name=reg_theme,
                    font_family=reg_font_family,
                    figure_width=reg_fig_width,
                    figure_height=reg_fig_height,
                    show_grid=reg_show_grid,
                )
                st.pyplot(regression_residual_fig)

            except Exception as exc:
                st.error("Failed to estimate the regression model. Please check the selected variables and data.")
                st.exception(exc)
                regression_summary_table = pd.DataFrame()
                regression_coef_table = pd.DataFrame()
                robust_coef_table = pd.DataFrame()
                normality_df = pd.DataFrame()
                vif_df = pd.DataFrame()
                independent_corr_df = pd.DataFrame()
                hetero_summary_df = pd.DataFrame()
                glejser_coef_df = pd.DataFrame()
                autocorr_df = pd.DataFrame()
                outlier_df = pd.DataFrame()

# =====================================================
# TAB 6: EXPORT
# =====================================================

with tab_export:
    st.subheader("Export Charts and Excel Output")

    export_col_1, export_col_2, export_col_3 = st.columns(3)
    with export_col_1:
        dpi = st.selectbox("PNG resolution / DPI", [300, 600, 900, 1200, 1500], index=3)
    with export_col_2:
        transparent_background = st.checkbox("Transparent PNG background", value=False)
    with export_col_3:
        st.info("Default PNG resolution is 1200 DPI for publication-ready output.")

    st.markdown("#### Download PNG Charts")


    if 'chart_long_data' in locals() and isinstance(chart_long_data, pd.DataFrame) and not chart_long_data.empty and line_panel_variables:
        export_line_fig = create_panel_mean_line_chart(
            long_df=chart_long_data,
            x_col=line_x_col,
            panel_variables=line_panel_variables,
            title=line_title,
            subtitle=line_subtitle,
            x_label=line_x_label,
            y_label=line_y_label,
            figure_width=line_fig_width,
            figure_height=line_fig_height,
            panel_columns=line_panel_columns,
            share_y_axis=line_share_y_axis,
            theme_name=line_theme,
            font_family=line_font_family,
            line_colors=line_colors,
            line_width=line_width,
            show_markers=line_show_markers,
            marker_style=line_marker_style,
            marker_size=line_marker_size,
            show_value_labels=line_show_value_labels,
            value_label_decimal_digits=3,
            compact_labels=line_compact_labels,
            title_font_size=line_title_font_size,
            subtitle_font_size=line_subtitle_font_size,
            axis_font_size=line_axis_font_size,
            tick_font_size=line_tick_font_size,
            legend_font_size=line_legend_font_size,
            value_label_font_size=line_value_label_font_size,
            panel_title_font_size=line_panel_title_font_size,
            x_tick_rotation=line_x_tick_rotation,
            show_grid=line_show_grid,
            legend_position=line_legend_position,
        )
        line_png = fig_to_png_bytes(export_line_fig, dpi=dpi, transparent_background=transparent_background)
        st.download_button(
            label=f"⬇️ Download Multi-Panel Line Chart PNG ({dpi} DPI)",
            data=line_png,
            file_name="statcal_online_idx_banking_mean_line_panel_chart.png",
            mime="image/png",
        )
        plt.close(export_line_fig)


    if 'regression_model' in locals() and 'regression_residual_fig' in locals() and not regression_summary_table.empty:
        export_reg_fig = create_regression_diagnostic_figure(
            regression_model,
            theme_name=reg_theme,
            font_family=reg_font_family,
            figure_width=reg_fig_width,
            figure_height=reg_fig_height,
            show_grid=reg_show_grid,
        )
        regression_png = fig_to_png_bytes(export_reg_fig, dpi=dpi, transparent_background=transparent_background)
        st.download_button(
            label=f"⬇️ Download Regression Diagnostic Plot PNG ({dpi} DPI)",
            data=regression_png,
            file_name="statcal_online_idx_banking_regression_diagnostic_plot.png",
            mime="image/png",
        )
        plt.close(export_reg_fig)

    st.markdown("---")
    st.markdown("#### Download Excel Output")

    metadata_df = pd.DataFrame([{
        "Application": f"{APP_NAME} - {APP_TITLE}",
        "Website": WEBSITE_URL,
        "STATCAL ONLINE Page": STATCAL_ONLINE_URL,
        "Training Data URL": TRAINING_DATA_URL,
        "Source File": source_name,
        "Worksheet": sheet_name,
        "Rows Original": len(df),
        "Rows Filtered": len(filtered_df),
        "Selected Years": ", ".join([str(x) for x in selected_years]) if "Year" in columns else "",
        "Selected Ticker Codes": ", ".join([str(x) for x in selected_tickers]) if "Ticker Code" in columns else "",
        "Updated": APP_UPDATED,
    }])

    sheets_to_export = {
        "Metadata": metadata_df,
        "Filtered Data": filtered_df,
    }

    if 'descriptive_stats' in locals() and isinstance(descriptive_stats, pd.DataFrame) and not descriptive_stats.empty:
        sheets_to_export["Univariate Descriptive"] = descriptive_stats

    if 'grouped_descriptive_stats' in locals() and isinstance(grouped_descriptive_stats, pd.DataFrame) and not grouped_descriptive_stats.empty:
        sheets_to_export["Grouped Descriptive"] = grouped_descriptive_stats



    if 'chart_long_data' in locals() and isinstance(chart_long_data, pd.DataFrame) and not chart_long_data.empty:
        sheets_to_export["Line Chart Data"] = chart_long_data


    if 'regression_summary_table' in locals() and isinstance(regression_summary_table, pd.DataFrame) and not regression_summary_table.empty:
        sheets_to_export["Regression Model Fit"] = regression_summary_table

    if 'regression_coef_table' in locals() and isinstance(regression_coef_table, pd.DataFrame) and not regression_coef_table.empty:
        sheets_to_export["Regression Coefficients"] = regression_coef_table

    if 'robust_coef_table' in locals() and isinstance(robust_coef_table, pd.DataFrame) and not robust_coef_table.empty:
        sheets_to_export["Robust Coefficients"] = robust_coef_table

    if 'normality_df' in locals() and isinstance(normality_df, pd.DataFrame) and not normality_df.empty:
        sheets_to_export["Normality Tests"] = normality_df

    if 'vif_df' in locals() and isinstance(vif_df, pd.DataFrame) and not vif_df.empty:
        sheets_to_export["VIF Tolerance"] = vif_df

    if 'independent_corr_df' in locals() and isinstance(independent_corr_df, pd.DataFrame) and not independent_corr_df.empty:
        sheets_to_export["Regression Indep Corr"] = independent_corr_df.reset_index().rename(columns={"index": "Variable"})

    if 'hetero_summary_df' in locals() and isinstance(hetero_summary_df, pd.DataFrame) and not hetero_summary_df.empty:
        sheets_to_export["Heteroskedasticity"] = hetero_summary_df

    if 'glejser_coef_df' in locals() and isinstance(glejser_coef_df, pd.DataFrame) and not glejser_coef_df.empty:
        sheets_to_export["Glejser Coefficients"] = glejser_coef_df

    if 'autocorr_df' in locals() and isinstance(autocorr_df, pd.DataFrame) and not autocorr_df.empty:
        sheets_to_export["Autocorrelation"] = autocorr_df

    if 'outlier_df' in locals() and isinstance(outlier_df, pd.DataFrame) and not outlier_df.empty:
        sheets_to_export["Regression Outliers"] = outlier_df

    if 'regression_model_data' in locals() and isinstance(regression_model_data, pd.DataFrame) and not regression_model_data.empty:
        sheets_to_export["Regression Model Data"] = regression_model_data

    excel_bytes = dataframe_to_excel_bytes(sheets_to_export)
    st.download_button(
        label="⬇️ Download All Output Tables to Excel",
        data=excel_bytes,
        file_name="statcal_online_idx_banking_financial_data_analyzer_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
