
import math
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from scipy import stats
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="STATCAL ONLINE - IDX Banking Financial Data Analyzer",
    page_icon="🏦",
    layout="wide",
)


# =====================================================
# CONSTANTS
# =====================================================

APP_NAME = "STATCAL ONLINE"
APP_TITLE = "IDX Banking Financial Data Analyzer"
APP_UPDATED = "Last updated on June 19, 2026"
WEBSITE_URL = "https://statcal.com/"
STATCAL_ONLINE_URL = "https://statcal.com/statcal%20online.html"
TRAINING_DATA_URL = "https://drive.google.com/drive/folders/1s273Ad5FUElhzd5G16jWSBxbOtforzRR?usp=sharing"
SAMPLE_DATA_PATH = Path("data idx perbankan.xlsx")
LOGO_PATH = Path("logo_statcal.png")
BRAND_COLOR = "1F4E79"

FINANCIAL_URL_COL = "Financial Statement URL"

COLOR_PALETTES = {
    "Scopus Blue Orange": ["#1F4E79", "#E97132", "#70AD47", "#FFC000", "#7030A0", "#00A6A6", "#C00000", "#595959"],
    "Nature Publication": ["#1B7837", "#762A83", "#5AAE61", "#9970AB", "#A6DBA0", "#C2A5CF", "#00441B", "#40004B"],
    "Finance Professional": ["#0B3C49", "#3282B8", "#BBE1FA", "#F9A03F", "#D1495B", "#00798C", "#30638E", "#003D5B"],
    "Economics Journal": ["#003F5C", "#BC5090", "#FFA600", "#58508D", "#FF6361", "#2F4B7C", "#A05195", "#D45087"],
    "Black Gray Academic": ["#000000", "#404040", "#666666", "#808080", "#A6A6A6", "#BFBFBF", "#595959", "#262626"],
    "High Contrast Q1": ["#2166AC", "#B2182B", "#4D9221", "#762A83", "#D6604D", "#1B7837", "#4393C3", "#9970AB"],
}

THEMES = {
    "White Publication": {
        "figure_facecolor": "white",
        "axes_facecolor": "white",
        "text_color": "#111111",
        "grid_color": "#D9D9D9",
        "spine_color": "#222222",
    },
    "Light Gray Editorial": {
        "figure_facecolor": "#F7F7F7",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#111111",
        "grid_color": "#D0D0D0",
        "spine_color": "#333333",
    },
    "Warm Ivory Journal": {
        "figure_facecolor": "#FBF7EF",
        "axes_facecolor": "#FFFDF8",
        "text_color": "#1F1F1F",
        "grid_color": "#DDD4C4",
        "spine_color": "#3A3A3A",
    },
    "Cool Blue Scientific": {
        "figure_facecolor": "#F3F7FB",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#0B1F33",
        "grid_color": "#C8D6E5",
        "spine_color": "#1F4E79",
    },
    "Dark Navy Presentation": {
        "figure_facecolor": "#0B1320",
        "axes_facecolor": "#111C2E",
        "text_color": "#FFFFFF",
        "grid_color": "#3B4A5F",
        "spine_color": "#B8C7D9",
    },
}

MARKERS = ["o", "s", "^", "D", "P", "X", "v", "<", ">", "*"]


# =====================================================
# GENERAL HELPERS
# =====================================================

def safe_streamlit_image(image_path: Path, width: int = 220) -> None:
    try:
        st.image(str(image_path), width=width)
    except Exception:
        st.markdown("### STATCAL")


@st.cache_data(ttl=3600, max_entries=20)
def read_excel_file(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)


@st.cache_data(ttl=3600, max_entries=10)
def get_excel_sheet_names(file_bytes: bytes) -> List[str]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes))
    return excel_file.sheet_names


@st.cache_data(ttl=3600, max_entries=20)
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(col)).strip() for col in df.columns]
    df = df.dropna(how="all")
    unnamed_cols = [col for col in df.columns if str(col).lower().startswith("unnamed")]
    for col in unnamed_cols:
        if df[col].isna().all():
            df = df.drop(columns=[col])
    return df.reset_index(drop=True)


def make_arrow_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    for col in safe_df.columns:
        if safe_df[col].dtype == "object":
            safe_df[col] = safe_df[col].astype(str).replace({"nan": "", "None": "", "NaT": ""})
    return safe_df


def to_numeric_series(series: pd.Series) -> pd.Series:
    """Convert a pandas Series to numeric values. Supports commas, %, and K/M/B/T suffixes."""
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")

    text = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("−", "-", regex=False)
        .str.replace("—", "", regex=False)
        .str.strip()
    )

    text = text.replace({
        "": pd.NA,
        "nan": pd.NA,
        "None": pd.NA,
        "NaT": pd.NA,
        "-": pd.NA,
        "N/A": pd.NA,
        "NA": pd.NA,
        "n/a": pd.NA,
        "na": pd.NA,
    })

    def convert_value(value):
        if pd.isna(value):
            return np.nan
        value = str(value).strip()
        if not value:
            return np.nan

        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1].strip()

        multiplier = 1.0
        last_char = value[-1:].lower()
        if last_char == "k":
            multiplier = 1_000.0
            value = value[:-1]
        elif last_char == "m":
            multiplier = 1_000_000.0
            value = value[:-1]
        elif last_char == "b":
            multiplier = 1_000_000_000.0
            value = value[:-1]
        elif last_char == "t":
            multiplier = 1_000_000_000_000.0
            value = value[:-1]

        value = value.replace("%", "").strip()
        try:
            number = float(value) * multiplier
            return -number if negative else number
        except Exception:
            return np.nan

    return text.map(convert_value)


@st.cache_data(ttl=3600, max_entries=20)
def detect_numeric_columns(df: pd.DataFrame, min_valid_ratio: float = 0.45) -> List[str]:
    numeric_cols = []
    for col in df.columns:
        s = df[col]
        non_null = s.notna().sum()
        if non_null == 0:
            continue
        numeric_s = to_numeric_series(s)
        valid_ratio = numeric_s.notna().sum() / max(non_null, 1)
        if valid_ratio >= min_valid_ratio:
            numeric_cols.append(col)
    return numeric_cols


def sorted_unique_values(series: pd.Series) -> List:
    values = series.dropna().unique().tolist()
    try:
        return sorted(values)
    except Exception:
        return sorted(values, key=lambda x: str(x).lower())


def preferred_option(options: List[str], candidates: List[str], fallback_index: int = 0) -> int:
    for candidate in candidates:
        for idx, option in enumerate(options):
            if str(option).strip().lower() == candidate.lower():
                return idx
    return fallback_index


def default_numeric_columns(columns: List[str]) -> List[str]:
    excluded = {"year", "tahun"}
    return [col for col in columns if str(col).strip().lower() not in excluded][:5]


def get_theme(theme_name: str) -> Dict[str, str]:
    return THEMES.get(theme_name, THEMES["White Publication"])


def get_palette_color_list(palette_name: str) -> List[str]:
    return COLOR_PALETTES.get(palette_name, COLOR_PALETTES["Scopus Blue Orange"])


def format_numeric_label(value: float, decimal_digits: int, compact: bool = False) -> str:
    if pd.isna(value) or not np.isfinite(value):
        return ""
    if compact:
        abs_value = abs(value)
        if abs_value >= 1_000_000_000_000:
            return f"{value / 1_000_000_000_000:.{decimal_digits}f}T"
        if abs_value >= 1_000_000_000:
            return f"{value / 1_000_000_000:.{decimal_digits}f}B"
        if abs_value >= 1_000_000:
            return f"{value / 1_000_000:.{decimal_digits}f}M"
        if abs_value >= 1_000:
            return f"{value / 1_000:.{decimal_digits}f}K"
    return f"{value:.{decimal_digits}f}"


def format_x_labels(values: List) -> List[str]:
    labels = []
    for value in values:
        if pd.isna(value):
            labels.append("Missing")
        elif isinstance(value, pd.Timestamp):
            labels.append(value.strftime("%Y-%m-%d"))
        elif isinstance(value, (np.integer, int)):
            labels.append(str(int(value)))
        else:
            labels.append(str(value))
    return labels


# =====================================================
# STATISTICS
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def compute_descriptive_statistics(df: pd.DataFrame, numeric_cols: List[str], decimal_digits: int) -> pd.DataFrame:
    rows = []
    for col in numeric_cols:
        s = to_numeric_series(df[col]).dropna()
        rows.append({
            "Variable": col,
            "N": int(s.count()),
            "Minimum": s.min() if len(s) else np.nan,
            "Maximum": s.max() if len(s) else np.nan,
            "Mean": s.mean() if len(s) else np.nan,
            "Median": s.median() if len(s) else np.nan,
            "Standard Deviation": s.std(ddof=1) if len(s) > 1 else np.nan,
            "Variance": s.var(ddof=1) if len(s) > 1 else np.nan,
            "Skewness": s.skew() if len(s) > 2 else np.nan,
            "Kurtosis": s.kurt() if len(s) > 3 else np.nan,
        })
    result = pd.DataFrame(rows)
    numeric_output_cols = [c for c in result.columns if c not in ["Variable", "N"]]
    for col in numeric_output_cols:
        result[col] = result[col].round(decimal_digits)
    return result


@st.cache_data(ttl=3600, max_entries=20)
def compute_grouped_descriptive_statistics(
    df: pd.DataFrame,
    numeric_cols: List[str],
    group_cols: List[str],
    decimal_digits: int,
) -> pd.DataFrame:
    if not group_cols:
        return compute_descriptive_statistics(df, numeric_cols, decimal_digits)

    rows = []
    for numeric_col in numeric_cols:
        temp = df[group_cols].copy()
        temp["__value__"] = to_numeric_series(df[numeric_col])
        grouped = temp.groupby(group_cols, dropna=False)["__value__"]
        stats_df = grouped.agg(
            N="count",
            Minimum="min",
            Maximum="max",
            Mean="mean",
            Median="median",
            **{"Standard Deviation": "std"},
        ).reset_index()
        stats_df.insert(len(group_cols), "Variable", numeric_col)
        rows.append(stats_df)

    result = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    for col in ["Minimum", "Maximum", "Mean", "Median", "Standard Deviation"]:
        if col in result.columns:
            result[col] = result[col].round(decimal_digits)
    return result


@st.cache_data(ttl=3600, max_entries=20)
def compute_correlation_matrix(
    df: pd.DataFrame,
    numeric_cols: List[str],
    method: str,
    decimal_digits: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    numeric_df = pd.DataFrame({col: to_numeric_series(df[col]) for col in numeric_cols})
    corr = numeric_df.corr(method=method.lower())

    pvals = pd.DataFrame(np.nan, index=numeric_cols, columns=numeric_cols)
    for row_var in numeric_cols:
        for col_var in numeric_cols:
            x = numeric_df[row_var]
            y = numeric_df[col_var]
            valid = pd.concat([x, y], axis=1).dropna()
            if len(valid) < 3:
                continue
            try:
                if method.lower() == "pearson":
                    _, p = stats.pearsonr(valid.iloc[:, 0], valid.iloc[:, 1])
                else:
                    _, p = stats.spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
                pvals.loc[row_var, col_var] = p
            except Exception:
                pvals.loc[row_var, col_var] = np.nan

    return corr.round(decimal_digits), pvals.round(decimal_digits)


# =====================================================
# LINE CHART
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def build_mean_line_long_data(
    df: pd.DataFrame,
    x_col: str,
    numeric_cols: List[str],
    split_col: str,
    sort_x: bool,
) -> pd.DataFrame:
    rows = []
    for variable in numeric_cols:
        use_cols = [x_col, variable]
        use_split = split_col != "None" and split_col in df.columns
        if use_split:
            use_cols.append(split_col)

        temp = df[use_cols].copy()
        temp[variable] = to_numeric_series(temp[variable])
        temp = temp.dropna(subset=[x_col, variable])
        if temp.empty:
            continue

        if use_split:
            grouped = temp.groupby([x_col, split_col], dropna=False)[variable].mean().reset_index()
            grouped = grouped.rename(columns={variable: "Mean"})
            grouped["Split"] = grouped[split_col].fillna("Missing").astype(str)
        else:
            grouped = temp.groupby(x_col, dropna=False)[variable].mean().reset_index()
            grouped = grouped.rename(columns={variable: "Mean"})
            grouped["Split"] = variable

        grouped["Variable"] = variable
        rows.append(grouped[[x_col, "Variable", "Split", "Mean"]])

    if not rows:
        return pd.DataFrame(columns=[x_col, "Variable", "Split", "Mean"])

    result = pd.concat(rows, ignore_index=True)
    if sort_x:
        try:
            result = result.sort_values(["Variable", "Split", x_col]).reset_index(drop=True)
        except Exception:
            result["__x_sort__"] = result[x_col].astype(str)
            result = result.sort_values(["Variable", "Split", "__x_sort__"]).drop(columns="__x_sort__").reset_index(drop=True)
    return result


def pivot_for_variable(long_df: pd.DataFrame, x_col: str, variable: str) -> pd.DataFrame:
    temp = long_df[long_df["Variable"] == variable].copy()
    if temp.empty:
        return pd.DataFrame()
    pivot = temp.pivot_table(index=x_col, columns="Split", values="Mean", aggfunc="mean")
    try:
        pivot = pivot.sort_index()
    except Exception:
        ordered = sorted(pivot.index.tolist(), key=lambda x: str(x).lower())
        pivot = pivot.loc[ordered]
    return pivot


def apply_common_chart_style(ax, theme: Dict[str, str], show_grid: bool) -> None:
    ax.set_facecolor(theme["axes_facecolor"])
    ax.tick_params(axis="both", colors=theme["text_color"])
    ax.xaxis.label.set_color(theme["text_color"])
    ax.yaxis.label.set_color(theme["text_color"])
    ax.title.set_color(theme["text_color"])

    if show_grid:
        ax.grid(axis="both", linestyle="--", linewidth=0.6, alpha=0.45, color=theme["grid_color"])
        ax.set_axisbelow(True)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_color(theme["spine_color"])
    ax.spines["left"].set_color(theme["spine_color"])


def create_panel_mean_line_chart(
    long_df: pd.DataFrame,
    x_col: str,
    panel_variables: List[str],
    title: str,
    subtitle: str,
    x_label: str,
    y_label: str,
    figure_width: float,
    figure_height: float,
    panel_columns: int,
    share_y_axis: bool,
    theme_name: str,
    font_family: str,
    line_colors: Dict[str, str],
    line_width: float,
    show_markers: bool,
    marker_style: str,
    marker_size: int,
    show_value_labels: bool,
    value_label_decimal_digits: int,
    compact_labels: bool,
    title_font_size: int,
    subtitle_font_size: int,
    axis_font_size: int,
    tick_font_size: int,
    legend_font_size: int,
    value_label_font_size: int,
    panel_title_font_size: int,
    x_tick_rotation: int,
    show_grid: bool,
    legend_position: str,
) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)

    n_panels = len(panel_variables)
    ncols = max(1, min(panel_columns, n_panels))
    nrows = math.ceil(n_panels / ncols)

    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(figure_width, figure_height), sharey=share_y_axis, squeeze=False)
    fig.patch.set_facecolor(theme["figure_facecolor"])

    legend_handles = []
    legend_labels = []

    for panel_idx, variable in enumerate(panel_variables):
        row = panel_idx // ncols
        col = panel_idx % ncols
        ax = axes[row][col]

        pivot = pivot_for_variable(long_df, x_col=x_col, variable=variable)
        x_values = np.arange(len(pivot.index))
        x_labels = format_x_labels(pivot.index.tolist())

        for series_name in pivot.columns:
            y_values = pd.to_numeric(pivot[series_name], errors="coerce").to_numpy(dtype=float)
            color = line_colors.get(str(series_name), "#1F4E79")
            line, = ax.plot(
                x_values,
                y_values,
                label=str(series_name),
                color=color,
                linewidth=line_width,
                marker=marker_style if show_markers else None,
                markersize=marker_size if show_markers else 0,
                markeredgecolor="white",
                markeredgewidth=0.6,
            )
            if str(series_name) not in legend_labels:
                legend_handles.append(line)
                legend_labels.append(str(series_name))

            if show_value_labels:
                for x, y in zip(x_values, y_values):
                    if pd.isna(y) or not np.isfinite(y):
                        continue
                    ax.annotate(
                        format_numeric_label(y, value_label_decimal_digits, compact=compact_labels),
                        (x, y),
                        textcoords="offset points",
                        xytext=(0, 7),
                        ha="center",
                        fontsize=value_label_font_size,
                        color=theme["text_color"],
                    )

        ax.set_title(str(variable), fontsize=panel_title_font_size, fontweight="bold", color=theme["text_color"], pad=10)
        ax.set_xlabel(x_label, fontsize=axis_font_size, labelpad=8, color=theme["text_color"])
        ax.set_ylabel(y_label, fontsize=axis_font_size, labelpad=8, color=theme["text_color"])
        ax.set_xticks(x_values)
        ax.set_xticklabels(x_labels, rotation=x_tick_rotation, ha="right" if x_tick_rotation > 0 else "center", fontsize=tick_font_size, color=theme["text_color"])
        ax.tick_params(axis="y", labelsize=tick_font_size, colors=theme["text_color"])
        apply_common_chart_style(ax, theme, show_grid)

    for panel_idx in range(n_panels, nrows * ncols):
        row = panel_idx // ncols
        col = panel_idx % ncols
        axes[row][col].axis("off")

    fig.suptitle(title, fontsize=title_font_size, fontweight="bold", color=theme["text_color"], y=0.995)
    if subtitle.strip():
        fig.text(0.5, 0.965, subtitle, ha="center", va="top", fontsize=subtitle_font_size, color=theme["text_color"])

    if legend_handles:
        if legend_position == "Outside right":
            legend = fig.legend(legend_handles, legend_labels, loc="upper left", bbox_to_anchor=(1.005, 0.94), frameon=True, fontsize=legend_font_size)
            fig.tight_layout(rect=[0, 0, 0.86, 0.93])
        elif legend_position == "Bottom":
            legend = fig.legend(legend_handles, legend_labels, loc="lower center", bbox_to_anchor=(0.5, -0.005), ncol=min(4, len(legend_labels)), frameon=True, fontsize=legend_font_size)
            fig.tight_layout(rect=[0, 0.06, 1, 0.93])
        else:
            legend = fig.legend(legend_handles, legend_labels, loc="upper right", frameon=True, fontsize=legend_font_size)
            fig.tight_layout(rect=[0, 0, 1, 0.93])
        legend.get_frame().set_alpha(0.88)
        for text in legend.get_texts():
            text.set_color(theme["text_color"])
    else:
        fig.tight_layout(rect=[0, 0, 1, 0.93])

    return fig


# =====================================================
# SCATTERPLOT
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def build_scatter_correlation_table(
    df: pd.DataFrame,
    dependent_col: str,
    independent_cols: List[str],
    method: str,
    decimal_digits: int,
) -> pd.DataFrame:
    rows = []
    y = to_numeric_series(df[dependent_col])
    for x_col in independent_cols:
        x = to_numeric_series(df[x_col])
        valid = pd.concat([x, y], axis=1).dropna()
        if len(valid) < 3:
            corr_value = np.nan
            p_value = np.nan
        else:
            try:
                if method.lower() == "pearson":
                    corr_value, p_value = stats.pearsonr(valid.iloc[:, 0], valid.iloc[:, 1])
                else:
                    corr_value, p_value = stats.spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
            except Exception:
                corr_value = np.nan
                p_value = np.nan
        rows.append({
            "Dependent Variable": dependent_col,
            "Independent Variable": x_col,
            "Correlation Method": method,
            "N": len(valid),
            "Correlation": corr_value,
            "p-value": p_value,
        })
    result = pd.DataFrame(rows)
    for col in ["Correlation", "p-value"]:
        result[col] = result[col].round(decimal_digits)
    return result


def create_scatterplot_panel(
    df: pd.DataFrame,
    dependent_col: str,
    independent_cols: List[str],
    color_col: str,
    label_col: str,
    title: str,
    subtitle: str,
    figure_width: float,
    figure_height: float,
    panel_columns: int,
    theme_name: str,
    font_family: str,
    palette_name: str,
    marker_style: str,
    marker_size: int,
    marker_alpha: float,
    show_labels: bool,
    label_font_size: int,
    add_regression_line: bool,
    title_font_size: int,
    subtitle_font_size: int,
    axis_font_size: int,
    tick_font_size: int,
    panel_title_font_size: int,
    legend_font_size: int,
    show_grid: bool,
) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)
    palette = get_palette_color_list(palette_name)

    n_panels = len(independent_cols)
    ncols = max(1, min(panel_columns, n_panels))
    nrows = math.ceil(n_panels / ncols)

    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(figure_width, figure_height), squeeze=False)
    fig.patch.set_facecolor(theme["figure_facecolor"])

    use_color = color_col != "None" and color_col in df.columns
    if use_color:
        categories = sorted_unique_values(df[color_col])
        color_map = {str(cat): palette[idx % len(palette)] for idx, cat in enumerate(categories)}
    else:
        categories = ["All"]
        color_map = {"All": palette[0]}

    legend_handles = {}
    y_all = to_numeric_series(df[dependent_col])

    for idx, x_col in enumerate(independent_cols):
        row = idx // ncols
        col = idx % ncols
        ax = axes[row][col]
        x_all = to_numeric_series(df[x_col])

        plot_df = df.copy()
        plot_df["__x__"] = x_all
        plot_df["__y__"] = y_all
        plot_df = plot_df.dropna(subset=["__x__", "__y__"])

        if use_color:
            for cat in categories:
                temp = plot_df[plot_df[color_col].astype(str) == str(cat)]
                if temp.empty:
                    continue
                sc = ax.scatter(
                    temp["__x__"],
                    temp["__y__"],
                    s=marker_size,
                    marker=marker_style,
                    alpha=marker_alpha,
                    color=color_map[str(cat)],
                    edgecolor="white",
                    linewidth=0.5,
                    label=str(cat),
                )
                if str(cat) not in legend_handles:
                    legend_handles[str(cat)] = sc
        else:
            sc = ax.scatter(
                plot_df["__x__"],
                plot_df["__y__"],
                s=marker_size,
                marker=marker_style,
                alpha=marker_alpha,
                color=palette[0],
                edgecolor="white",
                linewidth=0.5,
                label="All",
            )
            legend_handles["All"] = sc

        if add_regression_line and len(plot_df) >= 2:
            try:
                x_values = plot_df["__x__"].to_numpy(dtype=float)
                y_values = plot_df["__y__"].to_numpy(dtype=float)
                slope, intercept = np.polyfit(x_values, y_values, 1)
                order = np.argsort(x_values)
                ax.plot(x_values[order], intercept + slope * x_values[order], color=theme["spine_color"], linewidth=1.2, linestyle="--")
            except Exception:
                pass

        if show_labels and label_col != "None" and label_col in df.columns:
            for _, r in plot_df.iterrows():
                ax.annotate(
                    str(r[label_col]),
                    (r["__x__"], r["__y__"]),
                    textcoords="offset points",
                    xytext=(4, 4),
                    ha="left",
                    fontsize=label_font_size,
                    color=theme["text_color"],
                )

        ax.set_title(f"{x_col} vs {dependent_col}", fontsize=panel_title_font_size, fontweight="bold", color=theme["text_color"], pad=10)
        ax.set_xlabel(x_col, fontsize=axis_font_size, color=theme["text_color"])
        ax.set_ylabel(dependent_col, fontsize=axis_font_size, color=theme["text_color"])
        ax.tick_params(axis="both", labelsize=tick_font_size, colors=theme["text_color"])
        apply_common_chart_style(ax, theme, show_grid)

    for idx in range(n_panels, nrows * ncols):
        row = idx // ncols
        col = idx % ncols
        axes[row][col].axis("off")

    fig.suptitle(title, fontsize=title_font_size, fontweight="bold", color=theme["text_color"], y=0.995)
    if subtitle.strip():
        fig.text(0.5, 0.965, subtitle, ha="center", va="top", fontsize=subtitle_font_size, color=theme["text_color"])

    if legend_handles and use_color:
        legend = fig.legend(
            list(legend_handles.values()),
            list(legend_handles.keys()),
            loc="upper left",
            bbox_to_anchor=(1.005, 0.94),
            frameon=True,
            fontsize=legend_font_size,
        )
        legend.get_frame().set_alpha(0.88)
        for text in legend.get_texts():
            text.set_color(theme["text_color"])
        fig.tight_layout(rect=[0, 0, 0.86, 0.93])
    else:
        fig.tight_layout(rect=[0, 0, 1, 0.93])

    return fig


# =====================================================
# CORRELATION HEATMAP
# =====================================================

def create_correlation_heatmap(
    corr: pd.DataFrame,
    title: str,
    figure_width: float,
    figure_height: float,
    cmap: str,
    theme_name: str,
    font_family: str,
    title_font_size: int,
    tick_font_size: int,
    value_font_size: int,
    show_values: bool,
) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)
    fig, ax = plt.subplots(figsize=(figure_width, figure_height))
    fig.patch.set_facecolor(theme["figure_facecolor"])
    ax.set_facecolor(theme["axes_facecolor"])

    im = ax.imshow(corr.values, cmap=cmap, vmin=-1, vmax=1)
    ax.set_title(title, fontsize=title_font_size, fontweight="bold", color=theme["text_color"], pad=12)

    ax.set_xticks(np.arange(len(corr.columns)))
    ax.set_yticks(np.arange(len(corr.index)))
    ax.set_xticklabels(corr.columns, rotation=45, ha="right", fontsize=tick_font_size, color=theme["text_color"])
    ax.set_yticklabels(corr.index, fontsize=tick_font_size, color=theme["text_color"])

    if show_values:
        for i in range(len(corr.index)):
            for j in range(len(corr.columns)):
                value = corr.iloc[i, j]
                color = "white" if abs(value) > 0.55 else "black"
                ax.text(j, i, f"{value:.2f}", ha="center", va="center", fontsize=value_font_size, color=color)

    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.ax.tick_params(labelsize=tick_font_size, colors=theme["text_color"])

    for spine in ax.spines.values():
        spine.set_color(theme["spine_color"])

    fig.tight_layout()
    return fig


# =====================================================
# EXPORT HELPERS
# =====================================================

def fig_to_png_bytes(fig, dpi: int, transparent_background: bool = False) -> bytes:
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="none" if transparent_background else fig.get_facecolor(),
        transparent=transparent_background,
    )
    buffer.seek(0)
    return buffer.getvalue()


def style_excel_workbook(workbook) -> None:
    header_fill = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 55)


def dataframe_to_excel_bytes(sheets: Dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            safe_sheet = re.sub(r"[^A-Za-z0-9 _-]", "", str(sheet_name)).strip()[:31]
            if not safe_sheet:
                safe_sheet = "Sheet"
            data.to_excel(writer, sheet_name=safe_sheet, index=False)
        style_excel_workbook(writer.book)
    output.seek(0)
    return output.getvalue()


# =====================================================
# HEADER
# =====================================================

with st.container():
    left, right = st.columns([1, 5])
    with left:
        if LOGO_PATH.exists():
            safe_streamlit_image(LOGO_PATH, width=210)
        else:
            st.markdown("### STATCAL")
    with right:
        st.title(APP_NAME)
        st.subheader(APP_TITLE)
        st.caption(
            "A Python Streamlit application for exploring IDX banking financial data through data filtering, "
            "descriptive statistics, grouped summaries, correlation matrices, publication-ready multi-panel line charts, "
            "scatterplot panels, and exportable analytical outputs."
        )

st.markdown(
    f"""
    **Website:** [{WEBSITE_URL}]({WEBSITE_URL})  
    **STATCAL ONLINE Page:** [{STATCAL_ONLINE_URL}]({STATCAL_ONLINE_URL})  
    **Training Data / Sample Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})  
    **{APP_UPDATED}**  
    **Purpose:** Explore IDX banking financial data by year and ticker code, create descriptive tables, correlation matrices, line charts, scatterplots, and export outputs for academic reporting.

    ---
    """
)


# =====================================================
# TABS
# =====================================================

tab_data, tab_uni, tab_group, tab_corr, tab_line, tab_scatter, tab_export = st.tabs([
    "1. Data & Filters",
    "2. Univariate Descriptive",
    "3. Grouped Descriptive",
    "4. Correlation Matrix",
    "5. Multi-Panel Line Chart",
    "6. Scatterplot",
    "7. Export Charts & Excel",
])


# =====================================================
# TAB 1: DATA INPUT AND FILTERS
# =====================================================

with tab_data:
    st.subheader("Data Input and Flexible Filters")
    st.write("Upload an Excel dataset or use the included sample dataset: **data idx perbankan.xlsx**.")
    st.markdown(f"**Training Data / Sample Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})")
    st.caption("Download the sample dataset from Google Drive, then upload it using the file uploader below.")

    uploaded_file = st.file_uploader(
        "Upload Excel file",
        type=["xlsx", "xls"],
        help="Upload financial data in Excel format.",
    )

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        source_name = uploaded_file.name
    elif SAMPLE_DATA_PATH.exists():
        file_bytes = SAMPLE_DATA_PATH.read_bytes()
        source_name = SAMPLE_DATA_PATH.name
    else:
        st.info("Please upload an Excel file to start the analysis.")
        st.stop()

    try:
        sheet_names = get_excel_sheet_names(file_bytes)
        sheet_name = st.selectbox("Worksheet", sheet_names, index=0)
        raw_df = read_excel_file(file_bytes, sheet_name)
        df = clean_dataframe(raw_df)
    except Exception as exc:
        st.error("Failed to read the Excel file.")
        st.exception(exc)
        st.stop()

    if df.empty:
        st.error("The selected worksheet is empty.")
        st.stop()

    columns = df.columns.tolist()
    numeric_candidates = detect_numeric_columns(df)

    st.markdown("#### Filter Settings")
    filter_col_1, filter_col_2 = st.columns(2)

    if "Year" in columns:
        year_values = sorted_unique_values(df["Year"])
        with filter_col_1:
            selected_years = st.multiselect("Filter by Year", year_values, default=year_values, format_func=lambda x: str(x))
    else:
        selected_years = []

    if "Ticker Code" in columns:
        ticker_values = sorted_unique_values(df["Ticker Code"])
        with filter_col_2:
            selected_tickers = st.multiselect("Filter by Ticker Code", ticker_values, default=ticker_values, format_func=lambda x: str(x))
    else:
        selected_tickers = []

    filtered_df = df.copy()
    if "Year" in columns and selected_years:
        filtered_df = filtered_df[filtered_df["Year"].isin(selected_years)]
    if "Ticker Code" in columns and selected_tickers:
        filtered_df = filtered_df[filtered_df["Ticker Code"].isin(selected_tickers)]

    if filtered_df.empty:
        st.error("No rows remain after filtering. Please adjust the Year or Ticker Code filter.")
        st.stop()

    st.markdown("#### Dataset Preview")
    st.write(f"**Source:** {source_name} | **Worksheet:** {sheet_name}")
    st.dataframe(make_arrow_safe_dataframe(filtered_df), use_container_width=True)

    metric_1, metric_2, metric_3, metric_4 = st.columns(4)
    with metric_1:
        st.metric("Original rows", f"{len(df):,}")
    with metric_2:
        st.metric("Rows after filtering", f"{len(filtered_df):,}")
    with metric_3:
        st.metric("Columns", f"{len(df.columns):,}")
    with metric_4:
        st.metric("Detected numeric variables", f"{len(numeric_candidates):,}")

    if FINANCIAL_URL_COL in filtered_df.columns:
        with st.expander("Financial Statement URL List", expanded=False):
            url_df = filtered_df[["Ticker Code", "Company Name", "Year", FINANCIAL_URL_COL]].copy() if all(c in filtered_df.columns for c in ["Ticker Code", "Company Name", "Year"]) else filtered_df[[FINANCIAL_URL_COL]].copy()
            st.dataframe(make_arrow_safe_dataframe(url_df), use_container_width=True)

    with st.expander("Detected numeric variables", expanded=False):
        st.write(numeric_candidates)


# =====================================================
# TAB 2: UNIVARIATE DESCRIPTIVE
# =====================================================

with tab_uni:
    st.subheader("Univariate Descriptive Statistics")

    if not numeric_candidates:
        st.warning("No numeric variables were detected.")
        st.stop()

    uni_col_1, uni_col_2 = st.columns([3, 1])
    with uni_col_1:
        uni_numeric_cols = st.multiselect(
            "Select numeric variables for univariate descriptive statistics",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="uni_numeric_cols",
        )
    with uni_col_2:
        uni_decimal_digits = st.slider("Decimal digits", 0, 8, 3, 1, key="uni_decimal_digits")

    if not uni_numeric_cols:
        st.warning("Please select at least one numeric variable.")
    else:
        descriptive_stats = compute_descriptive_statistics(filtered_df, uni_numeric_cols, uni_decimal_digits)
        st.dataframe(make_arrow_safe_dataframe(descriptive_stats), use_container_width=True)


# =====================================================
# TAB 3: GROUPED DESCRIPTIVE
# =====================================================

with tab_group:
    st.subheader("Grouped Descriptive Statistics")

    group_col_1, group_col_2 = st.columns([3, 2])
    with group_col_1:
        group_numeric_cols = st.multiselect(
            "Select numeric variables for grouped descriptive statistics",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="group_numeric_cols",
        )
    with group_col_2:
        default_group = [c for c in ["Year", "Ticker Code"] if c in columns]
        group_cols = st.multiselect(
            "Group by category variables",
            columns,
            default=default_group,
            key="group_cols",
        )

    group_decimal_digits = st.slider("Decimal digits for grouped statistics", 0, 8, 3, 1, key="group_decimal_digits")

    if not group_numeric_cols:
        st.warning("Please select at least one numeric variable.")
    elif not group_cols:
        st.info("Please select at least one group variable to generate grouped descriptive statistics.")
        grouped_descriptive_stats = pd.DataFrame()
    else:
        grouped_descriptive_stats = compute_grouped_descriptive_statistics(filtered_df, group_numeric_cols, group_cols, group_decimal_digits)
        st.dataframe(make_arrow_safe_dataframe(grouped_descriptive_stats), use_container_width=True)


# =====================================================
# TAB 4: CORRELATION MATRIX
# =====================================================

with tab_corr:
    st.subheader("Correlation Matrix")

    corr_settings_1, corr_settings_2, corr_settings_3 = st.columns([3, 1, 1])
    with corr_settings_1:
        corr_numeric_cols = st.multiselect(
            "Select numeric variables for correlation matrix",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="corr_numeric_cols",
        )
    with corr_settings_2:
        corr_method = st.selectbox("Correlation method", ["Pearson", "Spearman"], index=0)
    with corr_settings_3:
        corr_decimal_digits = st.slider("Decimal digits", 0, 8, 3, 1, key="corr_decimal_digits")

    heat_settings = st.expander("Heatmap Settings", expanded=True)
    with heat_settings:
        heat_col_1, heat_col_2, heat_col_3 = st.columns(3)
        with heat_col_1:
            heat_theme = st.selectbox("Heatmap theme", list(THEMES.keys()), index=0, key="heat_theme")
            heat_cmap = st.selectbox("Color map", ["coolwarm", "RdBu_r", "BrBG", "PiYG", "viridis", "plasma"], index=0)
        with heat_col_2:
            heat_width = st.slider("Heatmap width", 5.0, 18.0, 10.0, 0.5)
            heat_height = st.slider("Heatmap height", 4.0, 18.0, 8.0, 0.5)
        with heat_col_3:
            heat_show_values = st.checkbox("Show correlation values", value=True)
            heat_font_family = st.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2, key="heat_font")

        heat_title_font_size = st.slider("Heatmap title font size", 10, 36, 16, key="heat_title_font")
        heat_tick_font_size = st.slider("Heatmap tick font size", 6, 24, 9, key="heat_tick_font")
        heat_value_font_size = st.slider("Heatmap value font size", 6, 22, 8, key="heat_value_font")

    if len(corr_numeric_cols) < 2:
        st.warning("Please select at least two numeric variables for correlation analysis.")
        corr_matrix = pd.DataFrame()
        corr_pvalues = pd.DataFrame()
    else:
        corr_matrix, corr_pvalues = compute_correlation_matrix(filtered_df, corr_numeric_cols, corr_method, corr_decimal_digits)
        st.markdown("#### Correlation Coefficients")
        st.dataframe(corr_matrix, use_container_width=True)

        st.markdown("#### Correlation p-values")
        st.dataframe(corr_pvalues, use_container_width=True)

        heatmap_fig = create_correlation_heatmap(
            corr=corr_matrix,
            title=f"{corr_method} Correlation Matrix",
            figure_width=heat_width,
            figure_height=heat_height,
            cmap=heat_cmap,
            theme_name=heat_theme,
            font_family=heat_font_family,
            title_font_size=heat_title_font_size,
            tick_font_size=heat_tick_font_size,
            value_font_size=heat_value_font_size,
            show_values=heat_show_values,
        )
        st.pyplot(heatmap_fig)


# =====================================================
# TAB 5: MULTI-PANEL LINE CHART
# =====================================================

with tab_line:
    st.subheader("Multi-Panel Mean Line Chart")

    line_setting_1, line_setting_2, line_setting_3 = st.columns(3)
    with line_setting_1:
        x_index = preferred_option(columns, ["Year", "Tahun", "Date", "Time", "Period", "Periode"], fallback_index=0)
        line_x_col = st.selectbox("X-axis variable", columns, index=x_index, key="line_x_col")
        line_numeric_cols = st.multiselect(
            "Panel variables / numeric variables",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates)[:4],
            key="line_numeric_cols",
        )
    with line_setting_2:
        split_options = ["None"] + columns
        split_index = preferred_option(split_options, ["Ticker Code", "Company Name", "Year"], fallback_index=0)
        line_split_col = st.selectbox("Split lines by category", split_options, index=split_index, key="line_split_col")
        line_sort_x = st.checkbox("Sort X-axis", value=True, key="line_sort_x")
    with line_setting_3:
        line_max_panels = st.slider("Maximum panels", 1, 20, min(6, max(1, len(line_numeric_cols))), key="line_max_panels")
        line_panel_columns = st.slider("Number of panel columns", 1, 4, min(2, max(1, line_max_panels)), key="line_panel_cols")
        line_share_y_axis = st.checkbox("Share Y-axis across panels", value=False, key="line_share_y")

    style_expander = st.expander("Line Chart Settings", expanded=True)
    with style_expander:
        style_col_1, style_col_2, style_col_3 = st.columns(3)
        with style_col_1:
            line_font_family = st.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2, key="line_font")
            line_theme = st.selectbox("Chart background theme", list(THEMES.keys()), index=0, key="line_theme")
            line_palette = st.selectbox("Color palette", list(COLOR_PALETTES.keys()), index=0, key="line_palette")
        with style_col_2:
            line_fig_width = st.slider("Figure width", 6.0, 30.0, 15.0, 0.5, key="line_fig_width")
            line_fig_height = st.slider("Figure height", 4.0, 30.0, 10.0, 0.5, key="line_fig_height")
            line_width = st.slider("Line width", 0.5, 8.0, 2.4, 0.1, key="line_width")
        with style_col_3:
            line_show_markers = st.checkbox("Add markers to lines", value=True, key="line_markers")
            line_marker_style = st.selectbox("Marker style", MARKERS, index=0, key="line_marker_style")
            line_marker_size = st.slider("Marker size", 2, 18, 6, key="line_marker_size")

        line_title = st.text_area("Chart title", value="Mean Trend of IDX Banking Financial Variables", height=68, key="line_title")
        line_subtitle = st.text_input("Chart subtitle", value="Multi-panel chart based on mean values from filtered IDX banking data", key="line_subtitle")
        line_x_label = st.text_input("X-axis label", value=line_x_col, key="line_x_label")
        line_y_label = st.text_input("Y-axis label", value="Mean", key="line_y_label")

        line_show_value_labels = st.checkbox("Show mean values on chart", value=False, key="line_show_value")
        line_compact_labels = st.checkbox("Use compact value labels (K/M/B/T)", value=True, key="line_compact")
        line_x_tick_rotation = st.slider("X-axis label rotation", 0, 90, 0, 5, key="line_x_rot")
        line_show_grid = st.checkbox("Show grid", value=True, key="line_grid")
        line_legend_position = st.selectbox("Legend position", ["Outside right", "Bottom", "Best"], index=0, key="line_legend")

        line_title_font_size = st.slider("Main title font size", 10, 44, 18, key="line_title_font")
        line_subtitle_font_size = st.slider("Subtitle font size", 8, 30, 11, key="line_subtitle_font")
        line_panel_title_font_size = st.slider("Panel title font size", 8, 30, 13, key="line_panel_title_font")
        line_axis_font_size = st.slider("Axis label font size", 8, 30, 12, key="line_axis_font")
        line_tick_font_size = st.slider("Tick label font size", 6, 26, 10, key="line_tick_font")
        line_legend_font_size = st.slider("Legend font size", 6, 24, 9, key="line_legend_font")
        line_value_label_font_size = st.slider("Value label font size", 5, 22, 8, key="line_value_font")

    line_panel_variables = line_numeric_cols[:line_max_panels] if line_numeric_cols else []
    if not line_panel_variables:
        st.warning("Please select at least one numeric variable for the line chart.")
        chart_long_data = pd.DataFrame()
    else:
        chart_long_data = build_mean_line_long_data(
            filtered_df,
            x_col=line_x_col,
            numeric_cols=line_panel_variables,
            split_col=line_split_col,
            sort_x=line_sort_x,
        )

        if chart_long_data.empty:
            st.error("The mean line chart data is empty. Please adjust variables or filters.")
        else:
            line_series = sorted(chart_long_data["Split"].dropna().astype(str).unique().tolist(), key=lambda x: str(x).lower())
            palette = get_palette_color_list(line_palette)
            line_colors: Dict[str, str] = {}
            st.markdown("#### Line Colors")
            color_cols = st.columns(4)
            for idx, series_name in enumerate(line_series[:32]):
                with color_cols[idx % 4]:
                    line_colors[str(series_name)] = st.color_picker(f"{series_name}", palette[idx % len(palette)], key=f"line_color_{idx}_{series_name}")

            with st.expander("Mean line chart data preview", expanded=False):
                st.dataframe(make_arrow_safe_dataframe(chart_long_data.head(300)), use_container_width=True)

            line_fig = create_panel_mean_line_chart(
                long_df=chart_long_data,
                x_col=line_x_col,
                panel_variables=line_panel_variables,
                title=line_title,
                subtitle=line_subtitle,
                x_label=line_x_label,
                y_label=line_y_label,
                figure_width=line_fig_width,
                figure_height=line_fig_height,
                panel_columns=line_panel_columns,
                share_y_axis=line_share_y_axis,
                theme_name=line_theme,
                font_family=line_font_family,
                line_colors=line_colors,
                line_width=line_width,
                show_markers=line_show_markers,
                marker_style=line_marker_style,
                marker_size=line_marker_size,
                show_value_labels=line_show_value_labels,
                value_label_decimal_digits=3,
                compact_labels=line_compact_labels,
                title_font_size=line_title_font_size,
                subtitle_font_size=line_subtitle_font_size,
                axis_font_size=line_axis_font_size,
                tick_font_size=line_tick_font_size,
                legend_font_size=line_legend_font_size,
                value_label_font_size=line_value_label_font_size,
                panel_title_font_size=line_panel_title_font_size,
                x_tick_rotation=line_x_tick_rotation,
                show_grid=line_show_grid,
                legend_position=line_legend_position,
            )
            st.pyplot(line_fig)


# =====================================================
# TAB 6: SCATTERPLOT
# =====================================================

with tab_scatter:
    st.subheader("Scatterplot Panel: Dependent and Independent Variables")

    scatter_setting_1, scatter_setting_2, scatter_setting_3 = st.columns(3)
    with scatter_setting_1:
        dep_index = preferred_option(numeric_candidates, ["Net Income", "Return on Asset (ROA)", "Total Assets"], fallback_index=0)
        scatter_dependent = st.selectbox("Dependent variable", numeric_candidates, index=dep_index, key="scatter_dep")
    with scatter_setting_2:
        scatter_independent_options = [col for col in numeric_candidates if col != scatter_dependent]
        scatter_independent = st.multiselect(
            "Independent variables",
            scatter_independent_options,
            default=scatter_independent_options[:3],
            key="scatter_ind",
        )
    with scatter_setting_3:
        scatter_corr_method = st.selectbox("Correlation method", ["Pearson", "Spearman"], index=0, key="scatter_corr_method")
        scatter_decimal_digits = st.slider("Decimal digits", 0, 8, 3, 1, key="scatter_digits")

    scatter_style = st.expander("Scatterplot Settings", expanded=True)
    with scatter_style:
        sc_col_1, sc_col_2, sc_col_3 = st.columns(3)
        with sc_col_1:
            scatter_color_col = st.selectbox("Color points by category", ["None"] + columns, index=preferred_option(["None"] + columns, ["Ticker Code", "Year"], 0), key="scatter_color")
            scatter_label_col = st.selectbox("Point label variable", ["None"] + columns, index=preferred_option(["None"] + columns, ["Ticker Code", "Company Name"], 0), key="scatter_label")
            scatter_show_labels = st.checkbox("Show point labels", value=False, key="scatter_show_labels")
        with sc_col_2:
            scatter_theme = st.selectbox("Chart background theme", list(THEMES.keys()), index=0, key="scatter_theme")
            scatter_palette = st.selectbox("Color palette", list(COLOR_PALETTES.keys()), index=0, key="scatter_palette")
            scatter_font_family = st.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2, key="scatter_font")
        with sc_col_3:
            scatter_marker_style = st.selectbox("Marker style", MARKERS, index=0, key="scatter_marker")
            scatter_marker_size = st.slider("Marker size", 20, 500, 80, 10, key="scatter_marker_size")
            scatter_marker_alpha = st.slider("Marker alpha", 0.1, 1.0, 0.80, 0.05, key="scatter_alpha")

        scatter_add_regression = st.checkbox("Add simple regression line", value=True, key="scatter_reg_line")
        scatter_show_grid = st.checkbox("Show grid", value=True, key="scatter_grid")
        scatter_panel_columns = st.slider("Number of panel columns", 1, 4, 2, key="scatter_panel_cols")
        scatter_fig_width = st.slider("Figure width", 6.0, 30.0, 15.0, 0.5, key="scatter_fig_width")
        scatter_fig_height = st.slider("Figure height", 4.0, 30.0, 10.0, 0.5, key="scatter_fig_height")
        scatter_title = st.text_area("Chart title", value="Scatterplot Panel of IDX Banking Financial Variables", height=68, key="scatter_title")
        scatter_subtitle = st.text_input("Chart subtitle", value="Dependent and independent variable relationships from filtered data", key="scatter_subtitle")

        scatter_title_font_size = st.slider("Main title font size", 10, 44, 18, key="scatter_title_font")
        scatter_subtitle_font_size = st.slider("Subtitle font size", 8, 30, 11, key="scatter_subtitle_font")
        scatter_panel_title_font_size = st.slider("Panel title font size", 8, 30, 13, key="scatter_panel_title_font")
        scatter_axis_font_size = st.slider("Axis label font size", 8, 30, 12, key="scatter_axis_font")
        scatter_tick_font_size = st.slider("Tick label font size", 6, 26, 10, key="scatter_tick_font")
        scatter_legend_font_size = st.slider("Legend font size", 6, 24, 9, key="scatter_legend_font")
        scatter_label_font_size = st.slider("Point label font size", 5, 22, 8, key="scatter_label_font")

    if not scatter_independent:
        st.warning("Please select at least one independent variable.")
        scatter_corr_table = pd.DataFrame()
    else:
        scatter_corr_table = build_scatter_correlation_table(
            filtered_df,
            dependent_col=scatter_dependent,
            independent_cols=scatter_independent,
            method=scatter_corr_method,
            decimal_digits=scatter_decimal_digits,
        )
        st.markdown("#### Correlation Table")
        st.dataframe(make_arrow_safe_dataframe(scatter_corr_table), use_container_width=True)

        scatter_fig = create_scatterplot_panel(
            df=filtered_df,
            dependent_col=scatter_dependent,
            independent_cols=scatter_independent,
            color_col=scatter_color_col,
            label_col=scatter_label_col,
            title=scatter_title,
            subtitle=scatter_subtitle,
            figure_width=scatter_fig_width,
            figure_height=scatter_fig_height,
            panel_columns=scatter_panel_columns,
            theme_name=scatter_theme,
            font_family=scatter_font_family,
            palette_name=scatter_palette,
            marker_style=scatter_marker_style,
            marker_size=scatter_marker_size,
            marker_alpha=scatter_marker_alpha,
            show_labels=scatter_show_labels,
            label_font_size=scatter_label_font_size,
            add_regression_line=scatter_add_regression,
            title_font_size=scatter_title_font_size,
            subtitle_font_size=scatter_subtitle_font_size,
            axis_font_size=scatter_axis_font_size,
            tick_font_size=scatter_tick_font_size,
            panel_title_font_size=scatter_panel_title_font_size,
            legend_font_size=scatter_legend_font_size,
            show_grid=scatter_show_grid,
        )
        st.pyplot(scatter_fig)


# =====================================================
# TAB 7: EXPORT
# =====================================================

with tab_export:
    st.subheader("Export Charts and Excel Output")

    export_col_1, export_col_2, export_col_3 = st.columns(3)
    with export_col_1:
        dpi = st.selectbox("PNG resolution / DPI", [300, 600, 900, 1200, 1500], index=3)
    with export_col_2:
        transparent_background = st.checkbox("Transparent PNG background", value=False)
    with export_col_3:
        st.info("Default PNG resolution is 1200 DPI for publication-ready output.")

    st.markdown("#### Download PNG Charts")

    if 'heatmap_fig' in locals() and not corr_matrix.empty:
        export_heat_fig = create_correlation_heatmap(
            corr=corr_matrix,
            title=f"{corr_method} Correlation Matrix",
            figure_width=heat_width,
            figure_height=heat_height,
            cmap=heat_cmap,
            theme_name=heat_theme,
            font_family=heat_font_family,
            title_font_size=heat_title_font_size,
            tick_font_size=heat_tick_font_size,
            value_font_size=heat_value_font_size,
            show_values=heat_show_values,
        )
        heat_png = fig_to_png_bytes(export_heat_fig, dpi=dpi, transparent_background=transparent_background)
        st.download_button(
            label=f"⬇️ Download Correlation Heatmap PNG ({dpi} DPI)",
            data=heat_png,
            file_name="statcal_online_idx_banking_correlation_heatmap.png",
            mime="image/png",
        )
        plt.close(export_heat_fig)

    if 'chart_long_data' in locals() and isinstance(chart_long_data, pd.DataFrame) and not chart_long_data.empty and line_panel_variables:
        export_line_fig = create_panel_mean_line_chart(
            long_df=chart_long_data,
            x_col=line_x_col,
            panel_variables=line_panel_variables,
            title=line_title,
            subtitle=line_subtitle,
            x_label=line_x_label,
            y_label=line_y_label,
            figure_width=line_fig_width,
            figure_height=line_fig_height,
            panel_columns=line_panel_columns,
            share_y_axis=line_share_y_axis,
            theme_name=line_theme,
            font_family=line_font_family,
            line_colors=line_colors,
            line_width=line_width,
            show_markers=line_show_markers,
            marker_style=line_marker_style,
            marker_size=line_marker_size,
            show_value_labels=line_show_value_labels,
            value_label_decimal_digits=3,
            compact_labels=line_compact_labels,
            title_font_size=line_title_font_size,
            subtitle_font_size=line_subtitle_font_size,
            axis_font_size=line_axis_font_size,
            tick_font_size=line_tick_font_size,
            legend_font_size=line_legend_font_size,
            value_label_font_size=line_value_label_font_size,
            panel_title_font_size=line_panel_title_font_size,
            x_tick_rotation=line_x_tick_rotation,
            show_grid=line_show_grid,
            legend_position=line_legend_position,
        )
        line_png = fig_to_png_bytes(export_line_fig, dpi=dpi, transparent_background=transparent_background)
        st.download_button(
            label=f"⬇️ Download Multi-Panel Line Chart PNG ({dpi} DPI)",
            data=line_png,
            file_name="statcal_online_idx_banking_mean_line_panel_chart.png",
            mime="image/png",
        )
        plt.close(export_line_fig)

    if 'scatter_corr_table' in locals() and isinstance(scatter_corr_table, pd.DataFrame) and not scatter_corr_table.empty and scatter_independent:
        export_scatter_fig = create_scatterplot_panel(
            df=filtered_df,
            dependent_col=scatter_dependent,
            independent_cols=scatter_independent,
            color_col=scatter_color_col,
            label_col=scatter_label_col,
            title=scatter_title,
            subtitle=scatter_subtitle,
            figure_width=scatter_fig_width,
            figure_height=scatter_fig_height,
            panel_columns=scatter_panel_columns,
            theme_name=scatter_theme,
            font_family=scatter_font_family,
            palette_name=scatter_palette,
            marker_style=scatter_marker_style,
            marker_size=scatter_marker_size,
            marker_alpha=scatter_marker_alpha,
            show_labels=scatter_show_labels,
            label_font_size=scatter_label_font_size,
            add_regression_line=scatter_add_regression,
            title_font_size=scatter_title_font_size,
            subtitle_font_size=scatter_subtitle_font_size,
            axis_font_size=scatter_axis_font_size,
            tick_font_size=scatter_tick_font_size,
            panel_title_font_size=scatter_panel_title_font_size,
            legend_font_size=scatter_legend_font_size,
            show_grid=scatter_show_grid,
        )
        scatter_png = fig_to_png_bytes(export_scatter_fig, dpi=dpi, transparent_background=transparent_background)
        st.download_button(
            label=f"⬇️ Download Scatterplot Panel PNG ({dpi} DPI)",
            data=scatter_png,
            file_name="statcal_online_idx_banking_scatterplot_panel.png",
            mime="image/png",
        )
        plt.close(export_scatter_fig)

    st.markdown("---")
    st.markdown("#### Download Excel Output")

    metadata_df = pd.DataFrame([{
        "Application": f"{APP_NAME} - {APP_TITLE}",
        "Website": WEBSITE_URL,
        "STATCAL ONLINE Page": STATCAL_ONLINE_URL,
        "Training Data URL": TRAINING_DATA_URL,
        "Source File": source_name,
        "Worksheet": sheet_name,
        "Rows Original": len(df),
        "Rows Filtered": len(filtered_df),
        "Selected Years": ", ".join([str(x) for x in selected_years]) if "Year" in columns else "",
        "Selected Ticker Codes": ", ".join([str(x) for x in selected_tickers]) if "Ticker Code" in columns else "",
        "Updated": APP_UPDATED,
    }])

    sheets_to_export = {
        "Metadata": metadata_df,
        "Filtered Data": filtered_df,
    }

    if 'descriptive_stats' in locals() and isinstance(descriptive_stats, pd.DataFrame) and not descriptive_stats.empty:
        sheets_to_export["Univariate Descriptive"] = descriptive_stats

    if 'grouped_descriptive_stats' in locals() and isinstance(grouped_descriptive_stats, pd.DataFrame) and not grouped_descriptive_stats.empty:
        sheets_to_export["Grouped Descriptive"] = grouped_descriptive_stats

    if 'corr_matrix' in locals() and isinstance(corr_matrix, pd.DataFrame) and not corr_matrix.empty:
        sheets_to_export["Correlation Matrix"] = corr_matrix.reset_index().rename(columns={"index": "Variable"})

    if 'corr_pvalues' in locals() and isinstance(corr_pvalues, pd.DataFrame) and not corr_pvalues.empty:
        sheets_to_export["Correlation PValues"] = corr_pvalues.reset_index().rename(columns={"index": "Variable"})

    if 'chart_long_data' in locals() and isinstance(chart_long_data, pd.DataFrame) and not chart_long_data.empty:
        sheets_to_export["Line Chart Data"] = chart_long_data

    if 'scatter_corr_table' in locals() and isinstance(scatter_corr_table, pd.DataFrame) and not scatter_corr_table.empty:
        sheets_to_export["Scatter Correlation"] = scatter_corr_table

    excel_bytes = dataframe_to_excel_bytes(sheets_to_export)
    st.download_button(
        label="⬇️ Download All Output Tables to Excel",
        data=excel_bytes,
        file_name="statcal_online_idx_banking_financial_data_analyzer_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
