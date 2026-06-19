
import math
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from scipy import stats
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="STATCAL ONLINE - IDX Banking Financial Data Analyzer",
    page_icon="🏦",
    layout="wide",
)


# =====================================================
# CONSTANTS
# =====================================================

APP_NAME = "STATCAL ONLINE"
APP_TITLE = "IDX Banking Financial Descriptive and Correlation Analyzer"
APP_UPDATED = "Last updated on June 19, 2026"
WEBSITE_URL = "https://statcal.com/"
STATCAL_ONLINE_URL = "https://statcal.com/statcal%20online.html"
TRAINING_DATA_URL = "https://drive.google.com/drive/folders/1s273Ad5FUElhzd5G16jWSBxbOtforzRR?usp=sharing"
SAMPLE_DATA_PATH = Path("data idx perbankan.xlsx")
LOGO_PATH = Path("logo_statcal.png")
BRAND_COLOR = "1F4E79"

FINANCIAL_URL_COL = "Financial Statement URL"

COLOR_PALETTES = {
    "Scopus Blue Orange": ["#1F4E79", "#E97132", "#70AD47", "#FFC000", "#7030A0", "#00A6A6", "#C00000", "#595959"],
    "Nature Publication": ["#1B7837", "#762A83", "#5AAE61", "#9970AB", "#A6DBA0", "#C2A5CF", "#00441B", "#40004B"],
    "Finance Professional": ["#0B3C49", "#3282B8", "#BBE1FA", "#F9A03F", "#D1495B", "#00798C", "#30638E", "#003D5B"],
    "Economics Journal": ["#003F5C", "#BC5090", "#FFA600", "#58508D", "#FF6361", "#2F4B7C", "#A05195", "#D45087"],
    "Black Gray Academic": ["#000000", "#404040", "#666666", "#808080", "#A6A6A6", "#BFBFBF", "#595959", "#262626"],
    "High Contrast Q1": ["#2166AC", "#B2182B", "#4D9221", "#762A83", "#D6604D", "#1B7837", "#4393C3", "#9970AB"],
}

THEMES = {
    "White Publication": {
        "figure_facecolor": "white",
        "axes_facecolor": "white",
        "text_color": "#111111",
        "grid_color": "#D9D9D9",
        "spine_color": "#222222",
    },
    "Light Gray Editorial": {
        "figure_facecolor": "#F7F7F7",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#111111",
        "grid_color": "#D0D0D0",
        "spine_color": "#333333",
    },
    "Warm Ivory Journal": {
        "figure_facecolor": "#FBF7EF",
        "axes_facecolor": "#FFFDF8",
        "text_color": "#1F1F1F",
        "grid_color": "#DDD4C4",
        "spine_color": "#3A3A3A",
    },
    "Cool Blue Scientific": {
        "figure_facecolor": "#F3F7FB",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#0B1F33",
        "grid_color": "#C8D6E5",
        "spine_color": "#1F4E79",
    },
    "Dark Navy Presentation": {
        "figure_facecolor": "#0B1320",
        "axes_facecolor": "#111C2E",
        "text_color": "#FFFFFF",
        "grid_color": "#3B4A5F",
        "spine_color": "#B8C7D9",
    },
}

MARKERS = ["o", "s", "^", "D", "P", "X", "v", "<", ">", "*"]


# =====================================================
# GENERAL HELPERS
# =====================================================

def safe_streamlit_image(image_path: Path, width: int = 220) -> None:
    try:
        st.image(str(image_path), width=width)
    except Exception:
        st.markdown("### STATCAL")


@st.cache_data(ttl=3600, max_entries=20)
def read_excel_file(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)


@st.cache_data(ttl=3600, max_entries=10)
def get_excel_sheet_names(file_bytes: bytes) -> List[str]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes))
    return excel_file.sheet_names


@st.cache_data(ttl=3600, max_entries=20)
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(col)).strip() for col in df.columns]
    df = df.dropna(how="all")
    unnamed_cols = [col for col in df.columns if str(col).lower().startswith("unnamed")]
    for col in unnamed_cols:
        if df[col].isna().all():
            df = df.drop(columns=[col])
    return df.reset_index(drop=True)


def make_arrow_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    for col in safe_df.columns:
        if safe_df[col].dtype == "object":
            safe_df[col] = safe_df[col].astype(str).replace({"nan": "", "None": "", "NaT": ""})
    return safe_df


def to_numeric_series(series: pd.Series) -> pd.Series:
    """Convert a pandas Series to numeric values. Supports commas, %, and K/M/B/T suffixes."""
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")

    text = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("−", "-", regex=False)
        .str.replace("—", "", regex=False)
        .str.strip()
    )

    text = text.replace({
        "": pd.NA,
        "nan": pd.NA,
        "None": pd.NA,
        "NaT": pd.NA,
        "-": pd.NA,
        "N/A": pd.NA,
        "NA": pd.NA,
        "n/a": pd.NA,
        "na": pd.NA,
    })

    def convert_value(value):
        if pd.isna(value):
            return np.nan
        value = str(value).strip()
        if not value:
            return np.nan

        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1].strip()

        multiplier = 1.0
        last_char = value[-1:].lower()
        if last_char == "k":
            multiplier = 1_000.0
            value = value[:-1]
        elif last_char == "m":
            multiplier = 1_000_000.0
            value = value[:-1]
        elif last_char == "b":
            multiplier = 1_000_000_000.0
            value = value[:-1]
        elif last_char == "t":
            multiplier = 1_000_000_000_000.0
            value = value[:-1]

        value = value.replace("%", "").strip()
        try:
            number = float(value) * multiplier
            return -number if negative else number
        except Exception:
            return np.nan

    return text.map(convert_value)


@st.cache_data(ttl=3600, max_entries=20)
def detect_numeric_columns(df: pd.DataFrame, min_valid_ratio: float = 0.45) -> List[str]:
    numeric_cols = []
    for col in df.columns:
        s = df[col]
        non_null = s.notna().sum()
        if non_null == 0:
            continue
        numeric_s = to_numeric_series(s)
        valid_ratio = numeric_s.notna().sum() / max(non_null, 1)
        if valid_ratio >= min_valid_ratio:
            numeric_cols.append(col)
    return numeric_cols


def sorted_unique_values(series: pd.Series) -> List:
    values = series.dropna().unique().tolist()
    try:
        return sorted(values)
    except Exception:
        return sorted(values, key=lambda x: str(x).lower())


def preferred_option(options: List[str], candidates: List[str], fallback_index: int = 0) -> int:
    for candidate in candidates:
        for idx, option in enumerate(options):
            if str(option).strip().lower() == candidate.lower():
                return idx
    return fallback_index


def default_numeric_columns(columns: List[str]) -> List[str]:
    excluded = {"year", "tahun"}
    return [col for col in columns if str(col).strip().lower() not in excluded][:5]


def get_theme(theme_name: str) -> Dict[str, str]:
    return THEMES.get(theme_name, THEMES["White Publication"])


def get_palette_color_list(palette_name: str) -> List[str]:
    return COLOR_PALETTES.get(palette_name, COLOR_PALETTES["Scopus Blue Orange"])


def format_numeric_label(value: float, decimal_digits: int, compact: bool = False) -> str:
    if pd.isna(value) or not np.isfinite(value):
        return ""
    if compact:
        abs_value = abs(value)
        if abs_value >= 1_000_000_000_000:
            return f"{value / 1_000_000_000_000:.{decimal_digits}f}T"
        if abs_value >= 1_000_000_000:
            return f"{value / 1_000_000_000:.{decimal_digits}f}B"
        if abs_value >= 1_000_000:
            return f"{value / 1_000_000:.{decimal_digits}f}M"
        if abs_value >= 1_000:
            return f"{value / 1_000:.{decimal_digits}f}K"
    return f"{value:.{decimal_digits}f}"


def format_x_labels(values: List) -> List[str]:
    labels = []
    for value in values:
        if pd.isna(value):
            labels.append("Missing")
        elif isinstance(value, pd.Timestamp):
            labels.append(value.strftime("%Y-%m-%d"))
        elif isinstance(value, (np.integer, int)):
            labels.append(str(int(value)))
        else:
            labels.append(str(value))
    return labels


# =====================================================
# STATISTICS
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def compute_descriptive_statistics(df: pd.DataFrame, numeric_cols: List[str], decimal_digits: int) -> pd.DataFrame:
    rows = []
    for col in numeric_cols:
        s = to_numeric_series(df[col]).dropna()
        rows.append({
            "Variable": col,
            "N": int(s.count()),
            "Minimum": s.min() if len(s) else np.nan,
            "Maximum": s.max() if len(s) else np.nan,
            "Mean": s.mean() if len(s) else np.nan,
            "Median": s.median() if len(s) else np.nan,
            "Standard Deviation": s.std(ddof=1) if len(s) > 1 else np.nan,
            "Variance": s.var(ddof=1) if len(s) > 1 else np.nan,
            "Skewness": s.skew() if len(s) > 2 else np.nan,
            "Kurtosis": s.kurt() if len(s) > 3 else np.nan,
        })
    result = pd.DataFrame(rows)
    numeric_output_cols = [c for c in result.columns if c not in ["Variable", "N"]]
    for col in numeric_output_cols:
        result[col] = result[col].round(decimal_digits)
    return result


@st.cache_data(ttl=3600, max_entries=20)
def compute_grouped_descriptive_statistics(
    df: pd.DataFrame,
    numeric_cols: List[str],
    group_cols: List[str],
    decimal_digits: int,
) -> pd.DataFrame:
    if not group_cols:
        return compute_descriptive_statistics(df, numeric_cols, decimal_digits)

    rows = []
    for numeric_col in numeric_cols:
        temp = df[group_cols].copy()
        temp["__value__"] = to_numeric_series(df[numeric_col])
        grouped = temp.groupby(group_cols, dropna=False)["__value__"]
        stats_df = grouped.agg(
            N="count",
            Minimum="min",
            Maximum="max",
            Mean="mean",
            Median="median",
            **{"Standard Deviation": "std"},
        ).reset_index()
        stats_df.insert(len(group_cols), "Variable", numeric_col)
        rows.append(stats_df)

    result = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    for col in ["Minimum", "Maximum", "Mean", "Median", "Standard Deviation"]:
        if col in result.columns:
            result[col] = result[col].round(decimal_digits)
    return result


@st.cache_data(ttl=3600, max_entries=20)
def compute_correlation_matrix(
    df: pd.DataFrame,
    numeric_cols: List[str],
    method: str,
    decimal_digits: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    numeric_df = pd.DataFrame({col: to_numeric_series(df[col]) for col in numeric_cols})
    corr = numeric_df.corr(method=method.lower())

    pvals = pd.DataFrame(np.nan, index=numeric_cols, columns=numeric_cols)
    for row_var in numeric_cols:
        for col_var in numeric_cols:
            x = numeric_df[row_var]
            y = numeric_df[col_var]
            valid = pd.concat([x, y], axis=1).dropna()
            if len(valid) < 3:
                continue
            try:
                if method.lower() == "pearson":
                    _, p = stats.pearsonr(valid.iloc[:, 0], valid.iloc[:, 1])
                else:
                    _, p = stats.spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
                pvals.loc[row_var, col_var] = p
            except Exception:
                pvals.loc[row_var, col_var] = np.nan

    return corr.round(decimal_digits), pvals.round(decimal_digits)


# =====================================================
# CHART STYLE HELPER
# =====================================================

def apply_common_chart_style(ax, theme: Dict[str, str], show_grid: bool) -> None:
    ax.set_facecolor(theme["axes_facecolor"])
    ax.tick_params(axis="both", colors=theme["text_color"])
    ax.xaxis.label.set_color(theme["text_color"])
    ax.yaxis.label.set_color(theme["text_color"])
    ax.title.set_color(theme["text_color"])

    if show_grid:
        ax.grid(axis="both", linestyle="--", linewidth=0.6, alpha=0.45, color=theme["grid_color"])
        ax.set_axisbelow(True)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_color(theme["spine_color"])
    ax.spines["left"].set_color(theme["spine_color"])



# CORRELATION HEATMAP
# =====================================================

def create_correlation_heatmap(
    corr: pd.DataFrame,
    title: str,
    figure_width: float,
    figure_height: float,
    cmap: str,
    theme_name: str,
    font_family: str,
    title_font_size: int,
    tick_font_size: int,
    value_font_size: int,
    show_values: bool,
) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)
    fig, ax = plt.subplots(figsize=(figure_width, figure_height))
    fig.patch.set_facecolor(theme["figure_facecolor"])
    ax.set_facecolor(theme["axes_facecolor"])

    im = ax.imshow(corr.values, cmap=cmap, vmin=-1, vmax=1)
    ax.set_title(title, fontsize=title_font_size, fontweight="bold", color=theme["text_color"], pad=12)

    ax.set_xticks(np.arange(len(corr.columns)))
    ax.set_yticks(np.arange(len(corr.index)))
    ax.set_xticklabels(corr.columns, rotation=45, ha="right", fontsize=tick_font_size, color=theme["text_color"])
    ax.set_yticklabels(corr.index, fontsize=tick_font_size, color=theme["text_color"])

    if show_values:
        for i in range(len(corr.index)):
            for j in range(len(corr.columns)):
                value = corr.iloc[i, j]
                color = "white" if abs(value) > 0.55 else "black"
                ax.text(j, i, f"{value:.2f}", ha="center", va="center", fontsize=value_font_size, color=color)

    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.ax.tick_params(labelsize=tick_font_size, colors=theme["text_color"])

    for spine in ax.spines.values():
        spine.set_color(theme["spine_color"])

    fig.tight_layout()
    return fig



# =====================================================
# EXPORT HELPERS
# =====================================================

def fig_to_png_bytes(fig, dpi: int, transparent_background: bool = False) -> bytes:
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="none" if transparent_background else fig.get_facecolor(),
        transparent=transparent_background,
    )
    buffer.seek(0)
    return buffer.getvalue()


def style_excel_workbook(workbook) -> None:
    header_fill = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 55)


def dataframe_to_excel_bytes(sheets: Dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            safe_sheet = re.sub(r"[^A-Za-z0-9 _-]", "", str(sheet_name)).strip()[:31]
            if not safe_sheet:
                safe_sheet = "Sheet"
            data.to_excel(writer, sheet_name=safe_sheet, index=False)
        style_excel_workbook(writer.book)
    output.seek(0)
    return output.getvalue()


# =====================================================
# HEADER
# =====================================================

with st.container():
    left, right = st.columns([1, 5])
    with left:
        if LOGO_PATH.exists():
            safe_streamlit_image(LOGO_PATH, width=210)
        else:
            st.markdown("### STATCAL")
    with right:
        st.title(APP_NAME)
        st.subheader(APP_TITLE)
        st.caption(
            "A Python Streamlit application for exploring IDX banking financial data through data filtering, "
            "univariate descriptive statistics, grouped summaries, correlation matrices, publication-ready heatmaps, "
            "and exportable analytical outputs."
        )

st.markdown(
    f"""
    **Website:** [{WEBSITE_URL}]({WEBSITE_URL})  
    **STATCAL ONLINE Page:** [{STATCAL_ONLINE_URL}]({STATCAL_ONLINE_URL})  
    **Data Source / Training Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})  
    **{APP_UPDATED}**  
    **Purpose:** Explore IDX banking financial data by year and ticker code, create descriptive tables, correlation matrices, publication-ready correlation heatmaps, and export outputs for academic reporting.

    ---
    """
)


# =====================================================
# TABS
# =====================================================

tab_data, tab_uni, tab_group, tab_corr, tab_export = st.tabs([
    "1. Data & Filters",
    "2. Univariate Descriptive",
    "3. Grouped Descriptive",
    "4. Correlation Matrix",
    "5. Export Charts & Excel",
])


# =====================================================
# TAB 1: DATA INPUT AND FILTERS
# =====================================================

with tab_data:
    st.subheader("Data Input and Flexible Filters")
    st.write("Upload an Excel dataset or use the included sample dataset: **data idx perbankan.xlsx**.")
    st.markdown(f"**Training Data / Sample Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})")

    uploaded_file = st.file_uploader(
        "Upload Excel file",
        type=["xlsx", "xls"],
        help="Upload financial data in Excel format.",
    )

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        source_name = uploaded_file.name
    elif SAMPLE_DATA_PATH.exists():
        file_bytes = SAMPLE_DATA_PATH.read_bytes()
        source_name = SAMPLE_DATA_PATH.name
    else:
        st.info("Please upload an Excel file to start the analysis.")
        st.stop()

    try:
        sheet_names = get_excel_sheet_names(file_bytes)
        sheet_name = st.selectbox("Worksheet", sheet_names, index=0)
        raw_df = read_excel_file(file_bytes, sheet_name)
        df = clean_dataframe(raw_df)
    except Exception as exc:
        st.error("Failed to read the Excel file.")
        st.exception(exc)
        st.stop()

    if df.empty:
        st.error("The selected worksheet is empty.")
        st.stop()

    columns = df.columns.tolist()
    numeric_candidates = detect_numeric_columns(df)

    st.markdown("#### Filter Settings")
    filter_col_1, filter_col_2 = st.columns(2)

    if "Year" in columns:
        year_values = sorted_unique_values(df["Year"])
        with filter_col_1:
            selected_years = st.multiselect("Filter by Year", year_values, default=year_values, format_func=lambda x: str(x))
    else:
        selected_years = []

    if "Ticker Code" in columns:
        ticker_values = sorted_unique_values(df["Ticker Code"])
        with filter_col_2:
            selected_tickers = st.multiselect("Filter by Ticker Code", ticker_values, default=ticker_values, format_func=lambda x: str(x))
    else:
        selected_tickers = []

    filtered_df = df.copy()
    if "Year" in columns and selected_years:
        filtered_df = filtered_df[filtered_df["Year"].isin(selected_years)]
    if "Ticker Code" in columns and selected_tickers:
        filtered_df = filtered_df[filtered_df["Ticker Code"].isin(selected_tickers)]

    if filtered_df.empty:
        st.error("No rows remain after filtering. Please adjust the Year or Ticker Code filter.")
        st.stop()

    st.markdown("#### Dataset Preview")
    st.write(f"**Source:** {source_name} | **Worksheet:** {sheet_name}")
    st.dataframe(make_arrow_safe_dataframe(filtered_df), use_container_width=True)

    metric_1, metric_2, metric_3, metric_4 = st.columns(4)
    with metric_1:
        st.metric("Original rows", f"{len(df):,}")
    with metric_2:
        st.metric("Rows after filtering", f"{len(filtered_df):,}")
    with metric_3:
        st.metric("Columns", f"{len(df.columns):,}")
    with metric_4:
        st.metric("Detected numeric variables", f"{len(numeric_candidates):,}")

    if FINANCIAL_URL_COL in filtered_df.columns:
        with st.expander("Financial Statement URL List", expanded=False):
            url_df = filtered_df[["Ticker Code", "Company Name", "Year", FINANCIAL_URL_COL]].copy() if all(c in filtered_df.columns for c in ["Ticker Code", "Company Name", "Year"]) else filtered_df[[FINANCIAL_URL_COL]].copy()
            st.dataframe(make_arrow_safe_dataframe(url_df), use_container_width=True)

    with st.expander("Detected numeric variables", expanded=False):
        st.write(numeric_candidates)


# =====================================================
# TAB 2: UNIVARIATE DESCRIPTIVE
# =====================================================

with tab_uni:
    st.subheader("Univariate Descriptive Statistics")

    if not numeric_candidates:
        st.warning("No numeric variables were detected.")
        st.stop()

    uni_col_1, uni_col_2 = st.columns([3, 1])
    with uni_col_1:
        uni_numeric_cols = st.multiselect(
            "Select numeric variables for univariate descriptive statistics",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="uni_numeric_cols",
        )
    with uni_col_2:
        uni_decimal_digits = st.slider("Decimal digits", 0, 8, 3, 1, key="uni_decimal_digits")

    if not uni_numeric_cols:
        st.warning("Please select at least one numeric variable.")
    else:
        descriptive_stats = compute_descriptive_statistics(filtered_df, uni_numeric_cols, uni_decimal_digits)
        st.dataframe(make_arrow_safe_dataframe(descriptive_stats), use_container_width=True)


# =====================================================
# TAB 3: GROUPED DESCRIPTIVE
# =====================================================

with tab_group:
    st.subheader("Grouped Descriptive Statistics")

    group_col_1, group_col_2 = st.columns([3, 2])
    with group_col_1:
        group_numeric_cols = st.multiselect(
            "Select numeric variables for grouped descriptive statistics",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="group_numeric_cols",
        )
    with group_col_2:
        default_group = [c for c in ["Year", "Ticker Code"] if c in columns]
        group_cols = st.multiselect(
            "Group by category variables",
            columns,
            default=default_group,
            key="group_cols",
        )

    group_decimal_digits = st.slider("Decimal digits for grouped statistics", 0, 8, 3, 1, key="group_decimal_digits")

    if not group_numeric_cols:
        st.warning("Please select at least one numeric variable.")
    elif not group_cols:
        st.info("Please select at least one group variable to generate grouped descriptive statistics.")
        grouped_descriptive_stats = pd.DataFrame()
    else:
        grouped_descriptive_stats = compute_grouped_descriptive_statistics(filtered_df, group_numeric_cols, group_cols, group_decimal_digits)
        st.dataframe(make_arrow_safe_dataframe(grouped_descriptive_stats), use_container_width=True)


# =====================================================
# TAB 4: CORRELATION MATRIX
# =====================================================

with tab_corr:
    st.subheader("Correlation Matrix")

    corr_settings_1, corr_settings_2, corr_settings_3 = st.columns([3, 1, 1])
    with corr_settings_1:
        corr_numeric_cols = st.multiselect(
            "Select numeric variables for correlation matrix",
            numeric_candidates,
            default=default_numeric_columns(numeric_candidates),
            key="corr_numeric_cols",
        )
    with corr_settings_2:
        corr_method = st.selectbox("Correlation method", ["Pearson", "Spearman"], index=0)
    with corr_settings_3:
        corr_decimal_digits = st.slider("Decimal digits", 0, 8, 3, 1, key="corr_decimal_digits")

    heat_settings = st.expander("Heatmap Settings", expanded=True)
    with heat_settings:
        heat_col_1, heat_col_2, heat_col_3 = st.columns(3)
        with heat_col_1:
            heat_theme = st.selectbox("Heatmap theme", list(THEMES.keys()), index=0, key="heat_theme")
            heat_cmap = st.selectbox("Color map", ["coolwarm", "RdBu_r", "BrBG", "PiYG", "viridis", "plasma"], index=0)
        with heat_col_2:
            heat_width = st.slider("Heatmap width", 5.0, 18.0, 10.0, 0.5)
            heat_height = st.slider("Heatmap height", 4.0, 18.0, 8.0, 0.5)
        with heat_col_3:
            heat_show_values = st.checkbox("Show correlation values", value=True)
            heat_font_family = st.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2, key="heat_font")

        heat_title_font_size = st.slider("Heatmap title font size", 10, 36, 16, key="heat_title_font")
        heat_tick_font_size = st.slider("Heatmap tick font size", 6, 24, 9, key="heat_tick_font")
        heat_value_font_size = st.slider("Heatmap value font size", 6, 22, 8, key="heat_value_font")

    if len(corr_numeric_cols) < 2:
        st.warning("Please select at least two numeric variables for correlation analysis.")
        corr_matrix = pd.DataFrame()
        corr_pvalues = pd.DataFrame()
    else:
        corr_matrix, corr_pvalues = compute_correlation_matrix(filtered_df, corr_numeric_cols, corr_method, corr_decimal_digits)
        st.markdown("#### Correlation Coefficients")
        st.dataframe(corr_matrix, use_container_width=True)

        st.markdown("#### Correlation p-values")
        st.dataframe(corr_pvalues, use_container_width=True)

        heatmap_fig = create_correlation_heatmap(
            corr=corr_matrix,
            title=f"{corr_method} Correlation Matrix",
            figure_width=heat_width,
            figure_height=heat_height,
            cmap=heat_cmap,
            theme_name=heat_theme,
            font_family=heat_font_family,
            title_font_size=heat_title_font_size,
            tick_font_size=heat_tick_font_size,
            value_font_size=heat_value_font_size,
            show_values=heat_show_values,
        )
        st.pyplot(heatmap_fig)


# =====================================================
# TAB 5: EXPORT
# =====================================================

with tab_export:
    st.subheader("Export Charts and Excel Output")

    export_col_1, export_col_2, export_col_3 = st.columns(3)
    with export_col_1:
        dpi = st.selectbox("PNG resolution / DPI", [300, 600, 900, 1200, 1500], index=3)
    with export_col_2:
        transparent_background = st.checkbox("Transparent PNG background", value=False)
    with export_col_3:
        st.info("Default PNG resolution is 1200 DPI for publication-ready output.")

    st.markdown("#### Download PNG Charts")

    if 'heatmap_fig' in locals() and not corr_matrix.empty:
        export_heat_fig = create_correlation_heatmap(
            corr=corr_matrix,
            title=f"{corr_method} Correlation Matrix",
            figure_width=heat_width,
            figure_height=heat_height,
            cmap=heat_cmap,
            theme_name=heat_theme,
            font_family=heat_font_family,
            title_font_size=heat_title_font_size,
            tick_font_size=heat_tick_font_size,
            value_font_size=heat_value_font_size,
            show_values=heat_show_values,
        )
        heat_png = fig_to_png_bytes(export_heat_fig, dpi=dpi, transparent_background=transparent_background)
        st.download_button(
            label=f"⬇️ Download Correlation Heatmap PNG ({dpi} DPI)",
            data=heat_png,
            file_name="statcal_online_idx_banking_correlation_heatmap.png",
            mime="image/png",
        )
        plt.close(export_heat_fig)


    st.markdown("---")
    st.markdown("#### Download Excel Output")

    metadata_df = pd.DataFrame([{
        "Application": f"{APP_NAME} - {APP_TITLE}",
        "Website": WEBSITE_URL,
        "STATCAL ONLINE Page": STATCAL_ONLINE_URL,
        "Training Data URL": TRAINING_DATA_URL,
        "Source File": source_name,
        "Worksheet": sheet_name,
        "Rows Original": len(df),
        "Rows Filtered": len(filtered_df),
        "Selected Years": ", ".join([str(x) for x in selected_years]) if "Year" in columns else "",
        "Selected Ticker Codes": ", ".join([str(x) for x in selected_tickers]) if "Ticker Code" in columns else "",
        "Updated": APP_UPDATED,
    }])

    sheets_to_export = {
        "Metadata": metadata_df,
        "Filtered Data": filtered_df,
    }

    if 'descriptive_stats' in locals() and isinstance(descriptive_stats, pd.DataFrame) and not descriptive_stats.empty:
        sheets_to_export["Univariate Descriptive"] = descriptive_stats

    if 'grouped_descriptive_stats' in locals() and isinstance(grouped_descriptive_stats, pd.DataFrame) and not grouped_descriptive_stats.empty:
        sheets_to_export["Grouped Descriptive"] = grouped_descriptive_stats

    if 'corr_matrix' in locals() and isinstance(corr_matrix, pd.DataFrame) and not corr_matrix.empty:
        sheets_to_export["Correlation Matrix"] = corr_matrix.reset_index().rename(columns={"index": "Variable"})

    if 'corr_pvalues' in locals() and isinstance(corr_pvalues, pd.DataFrame) and not corr_pvalues.empty:
        sheets_to_export["Correlation PValues"] = corr_pvalues.reset_index().rename(columns={"index": "Variable"})

    excel_bytes = dataframe_to_excel_bytes(sheets_to_export)
    st.download_button(
        label="⬇️ Download All Output Tables to Excel",
        data=excel_bytes,
        file_name="statcal_online_idx_banking_descriptive_correlation_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
